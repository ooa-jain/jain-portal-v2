from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, flash, abort, send_from_directory
from pymongo import MongoClient
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from bson import ObjectId
from dotenv import load_dotenv
import json, os, io, openpyxl, random, secrets, base64, re
from datetime import datetime, timedelta
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _parse_jwt_payload(token_str):
    try:
        parts = token_str.split('.')
        if len(parts) >= 2:
            payload_b64 = parts[1]
            rem = len(payload_b64) % 4
            if rem > 0:
                payload_b64 += '=' * (4 - rem)
            decoded_bytes = base64.urlsafe_b64decode(payload_b64)
            data = json.loads(decoded_bytes.decode('utf-8'))
            if isinstance(data, dict) and data.get('email'):
                return data
    except Exception as e:
        print("JWT payload decode error:", e)
    return None

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'jain_ooa_semreadiness_2025_secret')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

@app.after_request
def add_cache_control_headers(response):
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ── Firebase (Google sign-in) — public web config, sourced from .env with
#    the project's values as fallback so it works out of the box. ──
FIREBASE_CONFIG = {
    'apiKey': os.environ.get('FIREBASE_API_KEY', 'AIzaSyDNEpEpAfwT_pyDRjZB--n5nZUWZwLq0go'),
    'authDomain': os.environ.get('FIREBASE_AUTH_DOMAIN', 'jain-university-d4f8b.firebaseapp.com'),
    'projectId': os.environ.get('FIREBASE_PROJECT_ID', 'jain-university-d4f8b'),
    'storageBucket': os.environ.get('FIREBASE_STORAGE_BUCKET', 'jain-university-d4f8b.firebasestorage.app'),
    'messagingSenderId': os.environ.get('FIREBASE_MESSAGING_SENDER_ID', '386264535656'),
    'appId': os.environ.get('FIREBASE_APP_ID', '1:386264535656:web:32aa745d1776a0185a21dd'),
}

@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
def serve_sw():
    return send_from_directory('static', 'sw.js')

@app.route('/favicon.ico')
def serve_favicon():
    return send_from_directory('static/images', 'jgi-icon-192.png', mimetype='image/png')

# ── Google Search Console site verification ──
@app.route('/google72248b527c4b75b0.html')
def google_site_verification():
    return app.response_class(
        'google-site-verification: google72248b527c4b75b0.html',
        mimetype='text/html')

@app.route('/robots.txt')
def robots_txt():
    body = (
        "User-agent: *\n"
        "Allow: /\n"
        "Sitemap: https://jain-sarathi.juooa.cloud/sitemap.xml\n"
    )
    return app.response_class(body, mimetype='text/plain')

@app.route('/sitemap.xml')
def sitemap_xml():
    body = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        '  <url><loc>https://jain-sarathi.juooa.cloud/</loc>'
        '<changefreq>weekly</changefreq><priority>1.0</priority></url>\n'
        '</urlset>\n'
    )
    return app.response_class(body, mimetype='application/xml')

# MongoDB connection — sourced from .env (MONGO_URI + DB_NAME)
MONGO_URI = os.environ.get('MONGO_URI') or os.environ.get('MONGODB_URI') or 'mongodb://localhost:27017/'
DB_NAME = os.environ.get('DB_NAME', 'semreadiness')

try:
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=3000)
    # Ping to check if connection is active
    client.admin.command('ping')
    db = client[DB_NAME]
except Exception as mongo_err:
    print("Warning: Standard MongoDB connection failed:", mongo_err)
    print("Falling back to mongomock for local development execution.")
    import mongomock
    client = mongomock.MongoClient()
    db = client[DB_NAME]

submissions_col = db['submissions']
faculty_submissions_col = db['faculty_submissions']  # Faculty's individual checklist submissions (for both readiness & closure)
iea_col = db['iea_submissions']  # Innovation & Emerging Areas programme/course submissions
iea_edit_requests_col = db['iea_edit_requests']  # Requests to reopen a finalised IEA submission
users_col = db['users']
settings_col = db['settings']

ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin2023')

_raw_admin_emails = os.environ.get('ADMIN_EMAILS', 'santosh.ks@jainuniversity.ac.in,admin@jainuniversity.ac.in,admin@juooa.cloud')
ADMIN_EMAILS = set(e.strip().lower() for e in _raw_admin_emails.split(',') if e.strip())

def is_admin_email(email):
    if not email:
        return False
    email = email.strip().lower()
    if email in ADMIN_EMAILS:
        return True
    try:
        global_settings = settings_col.find_one({'_id': 'global'}) or {}
        dynamic_admins = set(e.lower() for e in global_settings.get('admin_emails', []))
        if email in dynamic_admins:
            return True
        user = users_col.find_one({'email': email})
        if user and (user.get('is_admin') is True or user.get('role') == 'admin'):
            return True
    except Exception:
        pass
    return False

DEPARTMENTS = [
    "Office of Academic Affairs",
    "Office of Academic",
    "Department of Computer Science and Engineering",
    "Department of Information Science and Engineering",
    "Department of Aerospace Engineering",
    "Department of Civil Engineering",
    "Department of Mechanical Engineering",
    "Department of Electrical and Electronics Engineering",
    "Department of Electronics and Communication Engineering",
    "Department of Food Technology",
    "Department of Humanities & Social Sciences",
    "Department of CERSSE",
    "Department of SSER",
    "Department of Jainology",
    "Department of Marine Science",
    "Department of Economics",
    "Department of Performing Arts and Cultural Studies",
    "Department of Languages",
    "Department of Journalism and Mass Communication",
    "Department of Law",
    "Department of Chemistry and Biochemistry",
    "Department of Biotechnology and Genetics",
    "Department of Microbiology and Botany",
    "Department of Data Analytics and Mathematical Science",
    "Department of Forensic Science",
    "Department of Physics and Electronics",
    "Department of Psychology and Allied Sciences",
    "Department of Allied Healthcare and Sciences",
    "Department of Computer Science and IT",
    "Department of Animation and Virtual Reality",
    "Department of Commerce",
    "Department of Management Studies",
    "Department of Design",
    "Department of Art and Design",
]

# ══════════════════════════════════════════════════════════════════
# READINESS — HOD CHECKLIST (HOD-only items, faculty items removed)
# ══════════════════════════════════════════════════════════════════
HOD_SECTIONS = [
    {"title": "Section 1: Curriculum & Course Matrix", "items": [
        {"id": "1", "text": "Previous semester's closing report insights incorporated"},
        {"id": "2", "text": "Course matrix verified (codes, credits, CBCS, LTPE) against BoS docs"},
        {"id": "3", "text": "Course matrix uploaded on ERP"},
        {"id": "4", "text": "Open Elective courses cross-checked with offering departments"},
        {"id": "5", "text": "Students given minimum 1 week to choose electives after orientation"},
        {"id": "6", "text": "Kochi / Online / ODL / other campus stakeholders included in planning"},
    ]},
    {"title": "Section 2: Faculty Allocation & Workload", "items": [
        {"id": "7", "text": "All courses have faculty assigned"},
        {"id": "8", "text": "Workload is fair and transparent across the department"},
        {"id": "9", "text": "Faculty informed of their course allocations formally"},
        {"id": "10", "text": "Faculty workload documented and assessed"},
    ]},
    {"title": "Section 3: TLEP Audit (HOD-level)", "items": [
        {"id": "12", "text": "HOD has reviewed / audited all TLEPs"},
        {"id": "13", "text": "Innovative pedagogies and assessments collated from TLEPs"},
        {"id": "14", "text": "Previous semester faculty feedback shared with faculty"},
    ]},
    {"title": "Section 4: Timetable & ERP", "items": [
        {"id": "15", "text": "Master timetable prepared"},
        {"id": "16", "text": "Master timetable uploaded on ERP"},
        {"id": "17", "text": "Course matrix created and approved in Faculty"},
    ]},
    {"title": "Section 5: LMS & Digital Readiness (HOD-level)", "items": [
        {"id": "19", "text": "All faculty trained on LMS platform"},
    ]},
    {"title": "Section 6: Students & Activities", "items": [
        {"id": "21", "text": "Co-curricular and extra-curricular activities planned and calendarised"},
        {"id": "22", "text": "Student Progression and Graduation report reviewed"},
        {"id": "23", "text": "Departmental Academic Achievement report prepared"},
    ]},
    {"title": "Section 7: Faculty Development & Research", "items": [
        {"id": "24", "text": "Faculty development / capacity building activities planned"},
        {"id": "25", "text": "Research activities (conferences, seminars) planned and calendarised"},
    ]},
    {"title": "Section 8: Infrastructure & Approvals", "items": [
        {"id": "26", "text": "Lab equipment, requirements and budget assessed"},
        {"id": "27", "text": "All necessary approvals obtained from University HO"},
        {"id": "28", "text": "Mid Sem Review Date"},
    ]},
]

# ══════════════════════════════════════════════════════════════════
# READINESS — FACULTY CHECKLIST (each faculty fills for their course)
# ══════════════════════════════════════════════════════════════════
FACULTY_SECTIONS = [
    {"title": "Section A: TLEP & Course Planning", "items": [
        {"id": "F1", "text": "TLEP submitted for the course"},
        {"id": "F2", "text": "Session plans prepared for entire semester"},
        {"id": "F3", "text": "Course objectives & outcomes mapped to PO/PSO"},
        {"id": "F4", "text": "Assessment plan finalized CIE/SEE components"},
    ]},
    {"title": "Section B: ERP & Documentation", "items": [
        {"id": "F5", "text": "Session plan uploaded on ERP"},
        {"id": "F6", "text": "Course material (notes/PPTs) prepared"},
        {"id": "F7", "text": "Reference textbooks & resources listed"},
    ]},
    {"title": "Section C: LMS & Digital Readiness", "items": [
        {"id": "F8", "text": "LMS course is live & accessible to students"},
        {"id": "F9", "text": "Week-1 content uploaded on LMS"},
        {"id": "F10", "text": "Discussion forums / assignments configured on LMS"},
    ]},
    {"title": "Section D: Pedagogy & Innovation", "items": [
        {"id": "F11", "text": "Innovative pedagogies planned (case studies, flipped classroom, etc.)"},
        {"id": "F12", "text": "Industry/practical examples integrated in course"},
    ]},
]

# ══════════════════════════════════════════════════════════════════
# CLOSURE — HOD REPORT SECTIONS (14 verticals, HOD-only narratives)
# ══════════════════════════════════════════════════════════════════
HOD_CLOSURE_SECTIONS = [
    {"id": "1",  "title": "Overview of the Semester",                                  "hint": "Brief narrative of how the semester progressed overall."},
    {"id": "2",  "title": "Objectives Planned — Were They Achieved?",                  "hint": "State the key objectives set at the beginning and whether each was met."},
    {"id": "3",  "title": "Adherence to Academic Calendar",                            "hint": "Highlight any deviations from the calendar and corrective actions taken."},
    {"id": "4",  "title": "Course Completion Summary",                                 "hint": "Summary of courses offered and completed. Flag any incomplete ones with reasons."},
    {"id": "5",  "title": "Student Performance Overview",                              "hint": "Average grades, top performers, notable academic achievements."},
    {"id": "6",  "title": "Faculty Contributions",                                     "hint": "Research published, conferences presented, awards received this semester."},
    {"id": "7",  "title": "Conferences, Workshops, Seminars & Guest Lectures",         "hint": "List all events conducted with dates and participants."},
    {"id": "8",  "title": "Department Events & Extra-curricular Activities",           "hint": "Overview of departmental events, fests, competitions etc."},
    {"id": "9",  "title": "Collaborations & Partnerships",                             "hint": "New MoUs or collaborations with Industry / Academia signed this semester."},
    {"id": "10", "title": "Academic Audits",                                           "hint": "Confirm that academic audits were conducted for all courses. Note any findings."},
    {"id": "11", "title": "CA Components — Entry & Approval",                          "hint": "Confirm CA marks are entered and approved in ERP for all courses."},
    {"id": "12", "title": "COLAB Files Review",                                        "hint": "Status of COLAB file review for this semester."},
    {"id": "13", "title": "COLAB — Previous Closing, Current Opening & Attainment",    "hint": "Confirm COLAB linkage: previous closing report reviewed, current opening report ready, CO attainment computed."},
    {"id": "14", "title": "HOD Additional Observations / Action Items for Next Semester","hint": "Any other noteworthy points or action items to carry forward."},
]

# ══════════════════════════════════════════════════════════════════
# CLOSURE — FACULTY CHECKLIST (each faculty fills for their course)
# Mirrors TEMPLATE_Closure_Checklist_-_Faculty_Odd_Semester.xlsx
# ══════════════════════════════════════════════════════════════════
FACULTY_CLOSURE_SECTIONS = [
    {"title": "Section A: ERP & Records", "items": [
        {"id": "FC1",  "text": "Attendance Entry in ERP with Analysis Report"},
        {"id": "FC2",  "text": "Continuous Assessment Marks Entry in ERP"},
        {"id": "FC3",  "text": "CA components mapped to COs / Bloom's Level"},
        {"id": "FC4",  "text": "Consolidated CA list available"},
        {"id": "FC5",  "text": "Assignment record maintained"},
    ]},
    {"title": "Section B: Syllabus & Teaching", "items": [
        {"id": "FC6",  "text": "Compliance to Teaching-Learning-Evaluation Plan (TLEP)"},
        {"id": "FC7",  "text": "Completion of Syllabus"},
        {"id": "FC8",  "text": "Innovative teaching methods adapted"},
    ]},
    {"title": "Section C: Learning Materials", "items": [
        {"id": "FC9",  "text": "Availability of Lecture Notes / PPTs"},
        {"id": "FC10", "text": "Sample Lecture Notes / PPT uploaded"},
        {"id": "FC11", "text": "Experiential Learning record with rubrics for Evaluation"},
        {"id": "FC12", "text": "Sample Experiential Learning activity with rubrics uploaded"},
    ]},
    {"title": "Section D: Evaluation", "items": [
        {"id": "FC13", "text": "Innovative Evaluation Strategy used"},
        {"id": "FC14", "text": "Availability of CO-PO (and PSO, if applicable) mapping"},
        {"id": "FC15", "text": "Course End Survey Conducted"},
    ]},
    {"title": "Section E: Student Support", "items": [
        {"id": "FC16", "text": "Slow learner / advance learner list prepared"},
        {"id": "FC17", "text": "Record of remedial classes maintained"},
        {"id": "FC18", "text": "Record of Extra lectures for Guided Self Study (GSS)"},
        {"id": "FC19", "text": "Lab Manuals available (wherever lab is a part of CA)"},
        {"id": "FC20", "text": "Result analysis with backlog list (course-wise)"},
    ]},
    {"title": "Section F: Mentoring & Projects", "items": [
        {"id": "FC21", "text": "Mentoring Report prepared"},
        {"id": "FC22", "text": "TD-PCL Report (batches, students, faculty, project title, progress)"},
        {"id": "FC23", "text": "Internship Report — Groups"},
    ]},
]

# ══════════════════════════════════════════════════════════════════
# IEA — PROGRAMMES & COURSES IN INNOVATION AND EMERGING AREAS
# (UGC circular: data collection AY 2022-23 to AY 2026-27)
# ══════════════════════════════════════════════════════════════════
IEA_SCHOOLS = {
    "School of Sciences": {
        "Department of Chemistry & Biochemistry": ["UG"],
        "Department of Computer Science & IT": ["UG"],
        "Department of Data Analytics & Mathematical Sciences": ["UG"],
        "Department of Forensic Science": ["UG"],
        "Department of Microbiology & Botany": ["UG"],
        "Department of Physics & Electronics": ["UG"],
        "Department of Psychology & Allied Science": ["UG"],
    },
    "School of Commerce": {
        "Department of Commerce": ["UG", "PG", "Doctoral"],
    },
    "School of Computer Science & Information Technology": {
        "Department of Animation & Virtual Reality": ["UG", "PG"],
        "Department of Computer Science & Information Technology": ["UG", "PG"],
    },
    "School of Engineering & Technology": {
        "Department of Civil Engineering": ["UG", "PG"],
        "Department of Computer Science & Engineering": ["UG", "PG"],
        "Department of Electrical & Electronics Engineering": ["UG", "PG"],
        "Department of Electronics & Communication Engineering": ["UG", "PG"],
        "Department of Food Technology": ["UG", "PG"],
        "Department of Information Science & Engineering": ["UG"],
        "Department of Mechanical Engineering": ["UG", "PG"],
    },
    "School of Humanities & Social Sciences": {
        "Department of Economics": ["UG", "PG"],
        "Department of Journalism & Mass Communication": ["UG", "PG"],
        "Department of Languages": ["UG", "PG"],
        "Department of Psychology & Allied Science": ["UG", "PG"],
        "Department of CERSSE": ["UG", "PG", "Doctoral"],
        "Department of Humanities & Social Sciences": ["UG", "PG", "Doctoral"],
    },
    "School of Design, Media & Creative Arts": {
        "Department of Art & Design": ["UG", "PG"],
        "Department of Design": ["UG", "PG"],
        "Department of Performing Arts & Cultural Studies": ["UG", "PG"],
    },
    "School of Law": {
        "Department of Law": ["Integrated UG", "UG", "PG"],
    },
    "School of Allied Healthcare & Sciences": {
        "Department of Allied Healthcare & Sciences": ["UG", "PG"],
    },
    "Center for Management Studies (CMS)": {
        "Department of Management Studies": ["UG"],
    },
    "CMS Business School": {
        "Department of Management": ["PG"],
    },
    "School of Aerospace Engineering": {
        "Department of Aerospace Engineering": ["UG", "PG"],
    },
    "School for Aviation & Aerospace Management": {
        "Department of Aviation & Aerospace Management": ["UG", "PG", "Executive Programme", "Executive Postgraduate"],
    },
    "School of Sports Science & Research": {
        "Department of Physical Education & Sports": ["UG"],
        "Department of Sports Science": ["PG"],
    },
}

IEA_YEARS = ['AY 2022-23', 'AY 2023-24', 'AY 2024-25', 'AY 2025-26', 'AY 2026-27']

IEA_EVIDENCE_TYPES = [
    'Academic Council Approval',
    'Curriculum & Syllabus',
    'Programme Brochure',
    'Industry MoUs',
    'Professional Body Recognition',
    'Assessment Rubrics',
    'Student Project Reports',
    'Relevant Web Links',
    'Google Drive link (shared with access to Office of Academics)'
]

IEA_SECTIONS = [
    {"key": "A", "color": "#C9A227", "title": "New Programmes Introduced",
     "sub": "For each new programme introduced in this academic year",
     "fields": [
         {"k": "name", "l": "New programme introduced (IN INNOVATIVE AND EMERGING AREAS)", "t": "text"},
         {"k": "area", "l": "Innovation / Emerging area addressed", "t": "text",
          "ph": "e.g., AI, Data Science, Cyber Security, FinTech, Sustainability, Industry 4.0/5.0..."},
         {"k": "newCourses", "l": "New courses introduced", "t": "textarea"},
         {"k": "uniqueFeatures", "l": "Unique programme features", "t": "textarea"},
         {"k": "collaborations", "l": "Industry / Academic collaborations", "t": "textarea"},
     ]},
    {"key": "B", "color": "#2F6F4E", "title": "Innovations in Existing Programmes & Courses",
     "sub": "Significant innovations, course-level enhancements, and curriculum enhancements",
     "fields": [
         {"k": "programme", "l": "Programme Name", "t": "text", "ph": "e.g., B.Tech CSE / M.Sc Data Science"},
         {"k": "courseCode", "l": "Course Code", "t": "text", "ph": "e.g., 22CSE301 / MAT102"},
         {"k": "courseName", "l": "Course Name", "t": "text", "ph": "e.g., Advanced AI & Machine Learning"},
         {"k": "curriculumRevisions", "l": "Curriculum revisions", "t": "textarea"},
         {"k": "newElectives", "l": "New electives / minors / majors / specialisations", "t": "textarea"},
         {"k": "interdisciplinary", "l": "Interdisciplinary pathways", "t": "textarea"},
         {"k": "experiential", "l": "Experiential learning", "t": "textarea"},
     ]},
    {"key": "C", "color": "#2C5F8A", "title": "Recognition by Global Professional Bodies",
     "sub": "Accreditations, benchmarking, and their impact",
     "fields": [
         {"k": "programme", "l": "Programme Name", "t": "text"},
         {"k": "accreditations", "l": "Accreditations", "t": "textarea"},
         {"k": "recognitions", "l": "Recognitions", "t": "textarea"},
         {"k": "certifications", "l": "Certifications", "t": "textarea"},
         {"k": "curriculumMapping", "l": "Curriculum mapping", "t": "textarea"},
         {"k": "benchmarking", "l": "International benchmarking", "t": "textarea"},
         {"k": "impact", "l": "Impact on student outcomes", "t": "textarea"},
     ]},
    {"key": "D", "color": "#A85C2C", "title": "Industry-Integrated Skilling",
     "sub": "Industry-linked courses, certifications, and work-integrated learning",
     "fields": [
         {"k": "programme", "l": "Programme Name", "t": "text"},
         {"k": "industryIntegrated", "l": "Industry-integrated courses", "t": "textarea"},
         {"k": "embeddedCerts", "l": "Embedded certifications", "t": "textarea"},
         {"k": "internships", "l": "Internships / Apprenticeships", "t": "textarea"},
         {"k": "skillModules", "l": "Skill-based modules", "t": "textarea"},
         {"k": "jointDelivery", "l": "Joint delivery with industry", "t": "textarea"},
         {"k": "industryAssessment", "l": "Industry assessment", "t": "textarea"},
     ]},
    {"key": "E", "color": "#6B4A8A", "title": "Competency-Based Learning",
     "sub": "OBE, competency design, capstones, and assessment reform",
     "fields": [
         {"k": "programme", "l": "Programme Name", "t": "text"},
         {"k": "obe", "l": "OBE implementation", "t": "textarea"},
         {"k": "competencyCurriculum", "l": "Competency-based curriculum", "t": "textarea"},
         {"k": "capstone", "l": "Capstone / Innovation projects", "t": "textarea"},
         {"k": "fieldImmersion", "l": "Field immersion / Clinical training", "t": "textarea"},
         {"k": "assessmentReforms", "l": "Assessment reforms", "t": "textarea"},
     ]},
    {"key": "F", "color": "#E11D48", "title": "Any Other Dimensions",
     "sub": "Any other dimensions the School may deem fit to highlight in support of Innovation and Emerging Areas",
     "fields": [
         {"k": "dimensionTitle", "l": "Dimension / Highlight", "t": "text", "ph": "e.g., Patents & IPR, Student startups, Faculty innovation, MoUs, Awards & Recognitions..."},
         {"k": "dimensionDetails", "l": "Details & Contribution to Innovation and Emerging Areas", "t": "textarea", "ph": "Describe this dimension and how it supports Innovation and Emerging Areas..."},
     ]},
]

# Definition shown to every user before starting an IEA submission.
IEA_DEFINITION = (
    "Definition: Programmes and Courses in Innovation and Emerging Areas are academic "
    "offerings that address new, evolving, or rapidly advancing domains of knowledge, "
    "skills, technologies, professions, and societal needs. Such offerings are "
    "characterized by their relevance to contemporary and future developments, "
    "interdisciplinary and multidisciplinary approaches, integration of research and "
    "innovation, responsiveness to industry and societal requirements, and alignment "
    "with global trends and national priorities. They may encompass emerging fields, "
    "transformative technologies, novel applications of existing disciplines, and "
    "innovative pedagogical approaches that prepare learners for future opportunities, "
    "challenges, and lifelong learning."
)

# Office that owns the IEA report — a finalised submission can only be reopened
# by this office, so every lock message points the department here.
IEA_OFFICE_EMAIL = 'officeofacademicaffairs@jainuniversity.ac.in'

# ── Mail Helper ──────────────────────────────────────────

def get_base_url():
    # If in a request context, get the actual current domain
    try:
        from flask import has_request_context, request
        if has_request_context():
            return request.host_url
    except Exception:
        pass

    # Fallback to configured BASE_URL or APP_URL in .env
    base_url = os.environ.get('BASE_URL') or os.environ.get('APP_URL')
    if base_url:
        if not base_url.endswith('/'):
            base_url += '/'
        return base_url

    return 'http://localhost:5000/'

def _get_smtp_config():
    """Return SMTP config from .env, falling back to Gmail defaults."""
    host = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
    port = int(os.environ.get('MAIL_PORT', 587))
    user = os.environ.get('EMAIL_USER', 'info.loginpanel@gmail.com')
    password = os.environ.get('EMAIL_PASS', 'wedbfepklgtwtugf')
    use_ssl = os.environ.get('MAIL_USE_SSL', 'False').lower() in ('true', '1', 'yes')
    use_tls = os.environ.get('MAIL_USE_TLS', 'True').lower() in ('true', '1', 'yes')
    return host, port, user, password, use_ssl, use_tls

def _send_email(to_email, subject, html_content):
    try:
        host, port, user, password, use_ssl, use_tls = _get_smtp_config()

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = user
        msg['To'] = to_email

        part = MIMEText(html_content, 'html')
        msg.attach(part)

        if use_ssl:
            server = smtplib.SMTP_SSL(host, port)
        else:
            server = smtplib.SMTP(host, port)
            if use_tls:
                server.starttls()

        server.login(user, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"SMTP Error: {e}")
        return False

def send_otp_email(to_email, otp):
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">You requested to reset your security passcode.</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Please use the following 4-digit OTP to proceed:</p>
          <div style="background: #f8f9fb; border: 2px dashed #0A2558; border-radius: 8px; padding: 15px; font-size: 32px; font-weight: bold; letter-spacing: 8px; color: #0A2558; margin-bottom: 30px;">
            {otp}
          </div>
          <p style="color: #8892aa; font-size: 12px;">If you did not request this, please ignore this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, 'Your Security Passcode OTP', html_content)

def send_admin_otp_email(to_email, otp):
    print(f"=== ADMIN OTP FOR {to_email}: {otp} ===", flush=True)
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal - Admin Access</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Admin login attempt detected.</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Please use the following 6-digit OTP to authorize admin dashboard access:</p>
          <div style="background: #f8f9fb; border: 2px dashed #0A2558; border-radius: 8px; padding: 15px; font-size: 32px; font-weight: bold; letter-spacing: 8px; color: #0A2558; margin-bottom: 30px;">
            {otp}
          </div>
          <p style="color: #8892aa; font-size: 12px;">This code is valid for 10 minutes. If you did not make this login attempt, please secure your credentials immediately.</p>
        </div>
      </body>
    </html>
    """
    success = _send_email(to_email, 'Admin Login OTP - Action Required', html_content)
    if not success:
        print(f"SMTP sending failed, bypassing to allow login for testing: OTP is {otp}", flush=True)
    return True



def send_faculty_submission_email(hod_email, faculty_name, course_name, form_type, hod_name):
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear {hod_name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Faculty member <strong>{faculty_name}</strong> has submitted their {form_type} checklist for the course <strong>{course_name}</strong>.</p>
          <p style="color: #8892aa; font-size: 12px;">You can log in to the portal to review this submission.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(hod_email, f'Faculty Submission Update: {course_name}', html_content)

def send_faculty_confirmation_email(faculty_email, faculty_name, course_name, form_type):
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear {faculty_name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Thank you for completing your checklist! Your submission for the course <strong>{course_name}</strong> ({form_type}) has been successfully logged.</p>
          <p style="color: #8892aa; font-size: 12px;">This is an automated confirmation of your submission.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(faculty_email, f'Confirmation: Submission Successful for {course_name}', html_content)

def send_deadline_reminder_email(to_email, name, module_name, deadline_str):
    login_url = get_base_url() + "login"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">Semester Readiness & Closure Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear HOD {name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">This is a friendly reminder that the submission window for <strong>{module_name}</strong> will close in <strong>2 days</strong>.</p>
          <p style="color: #ef4444; font-weight: bold; font-size: 16px; margin-bottom: 25px;">Official Deadline: {deadline_str.replace('T', ' ')}</p>
          <p style="color: #3d4460; font-size: 14px; margin-bottom: 25px;">Please log in to the portal as soon as possible to complete and finalize your checklist before access is closed.</p>
          <div style="margin-bottom: 30px;">
            <a href="{login_url}" style="background-color: #0A2558; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Go to Login Portal</a>
          </div>
          <p style="color: #8892aa; font-size: 12px;">This is an automated reminder. If you have already finalized your submission, please ignore this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, f'IMPORTANT Reminder: {module_name} Submission Deadline is in 2 Days!', html_content)

def send_final_hour_deadline_email(to_email, name, module_name, deadline_str):
    login_url = get_base_url() + "login"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #fef2f2; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); border-top: 6px solid #ef4444;">
          <h2 style="color: #b91c1c; margin-bottom: 20px;">⚠️ FINAL WARNING</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear HOD {name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">This is a critical reminder that the submission window for <strong>{module_name}</strong> will close in <strong>exactly 1 hour</strong>!</p>
          <p style="color: #ef4444; font-weight: bold; font-size: 18px; margin-bottom: 25px; background: #fee2e2; padding: 10px; border-radius: 6px; display: inline-block;">
            CLOSING DEADLINE: {deadline_str.replace('T', ' ')}
          </p>
          <p style="color: #3d4460; font-size: 14px; margin-bottom: 25px;">Late submissions will be automatically blocked. Please finalize and submit your form immediately to avoid lockout.</p>
          <div style="margin-bottom: 30px;">
            <a href="{login_url}" style="background-color: #ef4444; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Submit Form Now</a>
          </div>
          <p style="color: #8892aa; font-size: 12px;">This is an automated critical alert. If you have already finalized your submission, please ignore this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, f'⚠️ CRITICAL: {module_name} Submission Deadline is in 1 HOUR!', html_content)

def check_and_send_deadline_reminders():
    settings = get_global_settings()
    now = datetime.now()

    # Same 2-day / 1-hour warning cycle for every module (readiness, closure, IEA)
    for module_key in MODULE_ORDER:
        mod = MODULES[module_key]
        deadline = settings.get(mod['deadline_key'])
        if not deadline:
            continue
        try:
            dt = datetime.strptime(deadline, "%Y-%m-%dT%H:%M")
            diff_hours = (dt - now).total_seconds() / 3600.0
            date_str = dt.strftime("%d %B %Y")

            if diff_hours <= 0:
                continue

            for window, sender in (('2day', send_deadline_reminder_email),
                                   ('1hour', send_final_hour_deadline_email)):
                limit = 48 if window == '2day' else 1.0
                if diff_hours > limit:
                    continue
                for user in users_col.find():
                    email = user.get('email')
                    if not email:
                        continue
                    name = user.get('name') or 'HOD'
                    if has_module_submission(module_key, email):
                        continue
                    reminder_key = f"reminder_{module_key}_{window}_{deadline}_{email}"
                    if db['sent_reminders'].find_one({'_id': reminder_key}):
                        continue
                    print(f"Sending {window} deadline warning to HOD {name} ({email}) for {mod['label']}...")
                    sender(email, name, mod['label'], date_str)
                    db['sent_reminders'].insert_one({'_id': reminder_key,
                                                     'sent_at': datetime.utcnow().isoformat()})
        except Exception as e:
            print(f"Error checking {module_key} deadline reminders: {e}")

def start_deadline_scheduler():
    import threading
    import time
    def run_scheduler():
        print("Deadline reminder background scheduler started...")
        while True:
            try:
                check_and_send_deadline_reminders()
            except Exception as ex:
                print(f"Scheduler check error: {ex}")
            time.sleep(60)  # check every 60 seconds

    t = threading.Thread(target=run_scheduler, daemon=True)
    t.start()

# ── Helpers ───────────────────────────────────────────────

def is_deadline_passed(deadline_str):
    if not deadline_str:
        return False
    try:
        from datetime import datetime
        dt = datetime.strptime(deadline_str, "%Y-%m-%dT%H:%M")
        return datetime.now() > dt
    except Exception as e:
        print(f"Error parsing deadline {deadline_str}: {e}")
        return False
def get_global_settings():
    settings = settings_col.find_one({'_id': 'global'}) or {}
    settings.setdefault('readiness_enabled', True)
    settings.setdefault('readiness_deadline', '')
    settings.setdefault('closure_enabled', True)
    settings.setdefault('closure_deadline', '')
    settings.setdefault('iea_enabled', True)
    settings.setdefault('iea_deadline', '')
    settings.setdefault('enabled_years', ['2024-25', '2025-26', '2026-27', '2027-28'])
    settings.setdefault('enabled_semesters', ['Even', 'Odd'])
    return settings

# ── Module registry ───────────────────────────────────────
# Semester Readiness, Semester Closure and Innovation & Emerging Areas are the
# three HOD-facing modules. Everything that varies between them lives here so
# settings, deadlines, reminders and late-access all treat the three the same
# way instead of special-casing IEA.
MODULE_ORDER = ['readiness', 'closure', 'iea']

MODULES = {
    'readiness': {
        'key': 'readiness',
        'label': 'Semester Readiness',
        'enabled_key': 'readiness_enabled',
        'deadline_key': 'readiness_deadline',
        'override_field': 'readiness_access_override',
        'permission_field': 'can_readiness',
        'aliases': ['readiness'],
    },
    'closure': {
        'key': 'closure',
        'label': 'Semester Closure',
        'enabled_key': 'closure_enabled',
        'deadline_key': 'closure_deadline',
        'override_field': 'closure_access_override',
        'permission_field': 'can_closure',
        'aliases': ['closure'],
    },
    'iea': {
        'key': 'iea',
        'label': 'Innovation & Emerging Areas',
        'enabled_key': 'iea_enabled',
        'deadline_key': 'iea_deadline',
        'override_field': 'iea_access_override',
        'permission_field': 'can_iea',
        'aliases': ['iea', 'innovation', 'emerging'],
    },
}

def get_module(module_key):
    """Look up a module by its key ('readiness' / 'closure' / 'iea')."""
    return MODULES.get((module_key or '').strip().lower())

def module_from_label(label):
    """Map a stored access-request module label back to its registry entry."""
    text = (label or '').strip().lower()
    if not text:
        return None
    for key in MODULE_ORDER:
        mod = MODULES[key]
        if text == mod['label'].lower():
            return mod
    for key in MODULE_ORDER:
        mod = MODULES[key]
        if any(alias in text for alias in mod['aliases']):
            return mod
    return None

def has_module_submission(module_key, email):
    """True when this HOD has a finalised (non-draft) submission for the module."""
    if not email:
        return False
    if module_key == 'iea':
        return bool(iea_col.find_one({'submitterEmail': email, 'submitted': True}))
    return bool(submissions_col.find_one({
        'identity.submitterEmail': email,
        'form_type': module_key,
        '_draft': False
    }))

def get_module_deadline(settings, module_key):
    """Return (raw_deadline, 'DD Month YYYY') for a module; date is 'N/A' if unset/bad."""
    mod = get_module(module_key)
    if not mod:
        return '', 'N/A'
    raw = settings.get(mod['deadline_key']) or ''
    if not raw:
        return '', 'N/A'
    try:
        return raw, datetime.strptime(raw, "%Y-%m-%dT%H:%M").strftime("%d %B %Y")
    except Exception as ex:
        print(f"Error parsing {module_key} deadline {raw}: {ex}")
        return raw, 'N/A'

def _sections_for(form_type):
    """Return (hod_sections_or_report, faculty_sections) for the given form type."""
    if form_type == 'closure':
        return HOD_CLOSURE_SECTIONS, FACULTY_CLOSURE_SECTIONS
    return HOD_SECTIONS, FACULTY_SECTIONS

# ── Routes ──────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user_doc = users_col.find_one({'email': session['user_email']})
    dept = user_doc.get('department', '') if user_doc else ''
    user_email = session['user_email']

    # Query 3-module status for user's department
    readiness_sub = submissions_col.find_one({
        'form_type': 'readiness',
        '$or': [{'identity.dept': dept}, {'identity.submitterEmail': user_email}]
    }, sort=[('timestamp', -1)])

    closure_sub = submissions_col.find_one({
        'form_type': 'closure',
        '$or': [{'identity.dept': dept}, {'identity.submitterEmail': user_email}]
    }, sort=[('timestamp', -1)])

    iea_sub = iea_col.find_one({
        '$or': [{'department': dept}, {'submitterEmail': user_email}]
    }, sort=[('lastUpdated', -1)])

    dept_status = {
        'readiness': {
            'submitted': bool(readiness_sub),
            'date': readiness_sub.get('timestamp', '')[:10] if readiness_sub else '',
            'id': str(readiness_sub['_id']) if readiness_sub else ''
        },
        'closure': {
            'submitted': bool(closure_sub),
            'date': closure_sub.get('timestamp', '')[:10] if closure_sub else '',
            'id': str(closure_sub['_id']) if closure_sub else ''
        },
        'iea': {
            'submitted': bool(iea_sub and iea_sub.get('submitted')),
            'date': (iea_sub.get('submittedAt') or iea_sub.get('lastUpdated', ''))[:10] if iea_sub else '',
            'id': str(iea_sub['_id']) if iea_sub else ''
        }
    }

    is_admin = is_admin_email(user_email)
    user = {
        'email': user_email,
        'name': session.get('user_name', user_doc.get('name', 'User') if user_doc else 'User'),
        'department': dept,
        'picture': user_doc.get('picture', '') if user_doc else '',
        'auth_provider': user_doc.get('auth_provider', 'password') if user_doc else 'password',
        'is_admin': is_admin,
        'timeout_pref': user_doc.get('timeout_pref', 15) if user_doc else 15,
        'lock_enabled': user_doc.get('lock_enabled', False) if user_doc else False,
        'first_time_login': user_doc.get('first_time_login', True) if user_doc else False,
        'can_readiness': True if is_admin else user_doc.get('can_readiness', True) if user_doc else True,
        'can_closure': True if is_admin else user_doc.get('can_closure', True) if user_doc else True,
        'can_iea': True if is_admin else user_doc.get('can_iea', True) if user_doc else True,
    }
    settings = get_global_settings()
    return render_template('dashboard.html', user=user, dept_status=dept_status, settings=settings, departments=DEPARTMENTS)

@app.route('/analysis')
def hod_analysis_page():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user_doc = users_col.find_one({'email': session['user_email']})
    user = {
        'email': session['user_email'],
        'name': session['user_name'],
        'department': user_doc.get('department', '') if user_doc else '',
        'first_time_login': user_doc.get('first_time_login', True) if user_doc else False
    }
    return render_template('analysis.html', user=user)

@app.route('/api/hod/analysis')
def hod_analysis_api():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    
    email = session['user_email']
    hod_subs = list(submissions_col.find({'identity.submitterEmail': email}))
    parent_ids = [str(s['_id']) for s in hod_subs]
    parent_map = {str(s['_id']): s for s in hod_subs}
    
    fac_subs = list(faculty_submissions_col.find({'parent_submission_id': {'$in': parent_ids}}))
    
    results = []
    for f in fac_subs:
        parent = parent_map.get(f['parent_submission_id'])
        if not parent:
            continue
        results.append({
            '_id': str(f['_id']),
            'faculty_name': f.get('faculty_name', ''),
            'faculty_email': f.get('faculty_email', ''),
            'form_type': f.get('form_type', 'readiness'),
            'program': f.get('program', ''),
            'course_name': f.get('course_name', ''),
            'course_code': f.get('course_code', ''),
            'timestamp': f.get('timestamp', ''),
            'parent_semester': parent.get('identity', {}).get('semester', ''),
            'parent_ac_year': parent.get('identity', {}).get('acYear', ''),
            'parent_dept': parent.get('identity', {}).get('dept', '')
        })
        
    return jsonify({'ok': True, 'data': results})

def has_access_override(email, module_type):
    mod = get_module(module_type)
    if not mod:
        return False
    user = users_col.find_one({'email': email})
    if not user:
        return False
    return bool(user.get(mod['override_field'], False))

def get_user_context(user_email):
    if not user_email:
        return {}
    user_doc = users_col.find_one({'email': user_email}) or {}
    dept = user_doc.get('department', '')
    is_admin = is_admin_email(user_email)
    return {
        'email': user_email,
        'name': session.get('user_name', user_doc.get('name', 'User') if user_doc else 'User'),
        'department': dept,
        'picture': user_doc.get('picture', '') if user_doc else '',
        'auth_provider': user_doc.get('auth_provider', 'password') if user_doc else 'password',
        'is_admin': is_admin,
        'timeout_pref': user_doc.get('timeout_pref', 15) if user_doc else 15,
        'lock_enabled': user_doc.get('lock_enabled', False) if user_doc else False,
        'first_time_login': user_doc.get('first_time_login', True) if user_doc else False,
        'can_readiness': True if is_admin else user_doc.get('can_readiness', True) if user_doc else True,
        'can_closure': True if is_admin else user_doc.get('can_closure', True) if user_doc else True,
        'can_iea': True if is_admin else user_doc.get('can_iea', True) if user_doc else True,
    }

@app.route('/readiness')
def readiness():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = get_user_context(session['user_email'])
    settings = get_global_settings()
    if not session.get('admin') and (not settings.get('readiness_enabled', True) or not user.get('can_readiness', True)):
        return redirect('/')
    is_passed = is_deadline_passed(settings.get('readiness_deadline'))
    override = has_access_override(session['user_email'], 'readiness')
    req = db['access_requests'].find_one({'user_email': session['user_email'], 'module': 'Semester Readiness'})
    return render_template('form.html', 
                           hod_sections=HOD_SECTIONS, 
                           departments=DEPARTMENTS, 
                           user=user, 
                           view_mode='list', 
                           settings=settings,
                           is_deadline_passed=is_passed,
                           has_override=override,
                           extension_request=req)

@app.route('/readiness/form')
@app.route('/readiness/form/<sub_id>')
def readiness_form(sub_id=None):
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = get_user_context(session['user_email'])
    settings = get_global_settings()
    if not session.get('admin') and (not settings.get('readiness_enabled', True) or not user.get('can_readiness', True)):
        return redirect('/')
    is_passed = is_deadline_passed(settings.get('readiness_deadline'))
    override = has_access_override(session['user_email'], 'readiness')
    if not session.get('admin') and is_passed and not override:
        return render_template('deadline_passed.html', module='Semester Readiness', deadline=settings.get('readiness_deadline'))
    return render_template('form.html', hod_sections=HOD_SECTIONS, departments=DEPARTMENTS, user=user, view_mode='form', edit_id=sub_id, settings=settings)

@app.route('/closure')
def closure():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = get_user_context(session['user_email'])
    settings = get_global_settings()
    if not session.get('admin') and (not settings.get('closure_enabled', True) or not user.get('can_closure', True)):
        return redirect('/')
    is_passed = is_deadline_passed(settings.get('closure_deadline'))
    override = has_access_override(session['user_email'], 'closure')
    req = db['access_requests'].find_one({'user_email': session['user_email'], 'module': 'Semester Closure'})
    return render_template('closure.html',
                           hod_closure_sections=HOD_CLOSURE_SECTIONS,
                           departments=DEPARTMENTS,
                           user=user,
                           view_mode='list',
                           settings=settings,
                           is_deadline_passed=is_passed,
                           has_override=override,
                           extension_request=req)

@app.route('/closure/form')
@app.route('/closure/form/<sub_id>')
def closure_form(sub_id=None):
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = get_user_context(session['user_email'])
    settings = get_global_settings()
    if not session.get('admin') and (not settings.get('closure_enabled', True) or not user.get('can_closure', True)):
        return redirect('/')
    is_passed = is_deadline_passed(settings.get('closure_deadline'))
    override = has_access_override(session['user_email'], 'closure')
    if not session.get('admin') and is_passed and not override:
        return render_template('deadline_passed.html', module='Semester Closure', deadline=settings.get('closure_deadline'))
    return render_template('closure.html',
                           hod_closure_sections=HOD_CLOSURE_SECTIONS,
                           departments=DEPARTMENTS,
                           user=user,
                           view_mode='form',
                           edit_id=sub_id,
                           settings=settings)

# ═════════════════════════════════════════════════════════════
# IEA — Innovation & Emerging Areas (Programmes & Courses)
# ═════════════════════════════════════════════════════════════

def _iea_empty_years():
    return {y: {s['key']: [] for s in IEA_SECTIONS} for y in IEA_YEARS}

def _iea_norm_year(y):
    return str(y or '').replace('–', '-').replace('—', '-').strip()

def _iea_entry_is_filled(entry):
    """True only when an entry actually carries content.

    Opening a section — or adding an entry card and typing nothing into it —
    must never register as a submission. An entry counts once any field, any
    evidence type, link or remark has been filled in.
    """
    if not isinstance(entry, dict):
        return False
    skip = {'id', 'evidenceTypes', 'evidenceDetails'}
    for k, v in entry.items():
        if k in skip:
            continue
        if isinstance(v, str) and v.strip():
            return True
        if isinstance(v, (int, float)) and not isinstance(v, bool) and v:
            return True
    types = entry.get('evidenceTypes')
    if isinstance(types, list) and any(str(t or '').strip() for t in types):
        return True
    details = entry.get('evidenceDetails')
    if isinstance(details, dict):
        for d in details.values():
            if isinstance(d, dict) and any(str(v or '').strip() for v in d.values()):
                return True
    return False

def _iea_merge_years(parsed_years):
    """Normalise a submitted/stored years object against the master shape.

    Blank entries are dropped here, so every count, card, chart and export
    downstream sees only entries somebody genuinely filled in.
    """
    base = _iea_empty_years()
    if isinstance(parsed_years, dict):
        norm_map = {_iea_norm_year(k): v for k, v in parsed_years.items()}
        for y in IEA_YEARS:
            ydata = norm_map.get(_iea_norm_year(y))
            if not isinstance(ydata, dict):
                continue
            for s in IEA_SECTIONS:
                entries = ydata.get(s['key'])
                if isinstance(entries, list):
                    clean = []
                    for e in entries:
                        if not isinstance(e, dict):
                            continue
                        entry = {'evidenceTypes': [], 'evidenceLink': '', 'evidenceMissing': ''}
                        entry.update(e)
                        if not _iea_entry_is_filled(entry):
                            continue
                        clean.append(entry)
                    base[y][s['key']] = clean
    return base

def _iea_count_entries(years):
    """Total entries across every year/section of a (possibly partial) years blob."""
    n = 0
    if isinstance(years, dict):
        for _y, secs in years.items():
            if isinstance(secs, dict):
                for _k, lst in secs.items():
                    if isinstance(lst, list):
                        n += sum(1 for e in lst if _iea_entry_is_filled(e))
    return n

def _iea_version_timeline(doc):
    """Compact edit history — one row per submitted version, oldest first.

    Each stored history entry carries a full snapshot of that version's data.
    Those snapshots are far too heavy to hand to the browser, so the timeline
    ships only the version number, its date, its size and who filed it.
    """
    rows = []
    for h in (doc.get('history') or []):
        if not isinstance(h, dict):
            continue
        rows.append({
            'version': h.get('version', 1),
            'at': h.get('submittedAt') or h.get('lastUpdated') or '',
            'entries': _iea_count_entries(h.get('years')),
            'by': h.get('submitterName') or h.get('submitterEmail') or '',
        })
    if doc.get('submitted'):
        rows.append({
            'version': doc.get('version', 1),
            'at': doc.get('submittedAt') or doc.get('lastUpdated') or '',
            'entries': _iea_count_entries(doc.get('years')),
            'by': doc.get('submitterName') or doc.get('submitterEmail') or '',
            'current': True,
        })
    rows.sort(key=lambda r: r.get('version') or 0)
    return rows

def _iea_public_doc(doc):
    """Shape one IEA submission for a template/API payload."""
    doc['_id'] = str(doc['_id'])
    doc['years'] = _iea_merge_years(doc.get('years'))
    doc['versions'] = _iea_version_timeline(doc)
    # Derive when this department started filling for rows saved before the
    # field existed, so the timeline still has a starting point.
    if not doc.get('createdAt'):
        earliest = next((r['at'] for r in doc['versions'] if r.get('at')), '')
        doc['createdAt'] = earliest or doc.get('lastUpdated', '')
        doc['createdAtEstimated'] = True
    doc.pop('history', None)
    return doc

IEA_IST_OFFSET = timedelta(hours=5, minutes=30)   # stored timestamps are UTC

def _iea_parse_dt(value):
    """Parse a stored ISO timestamp (UTC, with or without a trailing Z)."""
    raw = str(value or '').strip().replace('Z', '')
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        try:
            return datetime.strptime(raw[:19], '%Y-%m-%dT%H:%M:%S')
        except ValueError:
            return None

def _iea_recent_activity(docs):
    """Who filed an IEA submission today or yesterday (India time), newest first.

    One row per department: the department, the person who filed it, when, how
    many entries it holds, and whether it is a final submission or a draft save.
    """
    now_ist = datetime.utcnow() + IEA_IST_OFFSET
    today, yesterday = now_ist.date(), (now_ist - timedelta(days=1)).date()
    rows = []
    for d in docs:
        # Whichever happened last: the final submission, or a later draft save.
        sub_dt = _iea_parse_dt(d.get('submittedAt')) if d.get('submitted') else None
        edit_dt = _iea_parse_dt(d.get('lastUpdated'))
        dt = max([x for x in (sub_dt, edit_dt) if x], default=None)
        if not dt:
            continue
        submitted = bool(sub_dt and dt == sub_dt)
        stamp = (d.get('submittedAt') if submitted else d.get('lastUpdated')) or ''
        local = dt + IEA_IST_OFFSET
        if local.date() == today:
            day = 'today'
        elif local.date() == yesterday:
            day = 'yesterday'
        else:
            continue
        version = d.get('version', 1)
        if submitted:
            status = '✓ Submitted' + (f' · v{version}' if version and version > 1 else '')
        elif d.get('submitted'):
            status = 'Edited after submitting'
        else:
            status = 'Draft saved'
        rows.append({
            'day': day,
            'school': d.get('school', ''),
            'department': d.get('department', ''),
            'level': d.get('level', ''),
            'person': d.get('submitterName') or d.get('submitterEmail') or 'Not recorded',
            'email': d.get('submitterEmail', ''),
            'entries': _iea_count_entries(d.get('years')),
            'submitted': bool(d.get('submitted')),
            'status': status,
            'version': version if d.get('submitted') else 0,
            'at': stamp,
            'time': local.strftime('%d %b, %I:%M %p').lstrip('0'),
            'sortKey': local.isoformat(),
        })
    rows.sort(key=lambda r: r['sortKey'], reverse=True)
    return rows

def _iea_calendar_events(docs):
    """Every dated IEA action, indexed by calendar day (India time).

    Feeds the date picker on the analysis page: pick a day and see which
    departments started filling, saved a draft, or submitted a version on it.
    """
    days = {}

    def add(stamp, dept_doc, kind, label, entries, person):
        dt = _iea_parse_dt(stamp)
        if not dt:
            return
        local = dt + IEA_IST_OFFSET
        key = local.strftime('%Y-%m-%d')
        days.setdefault(key, []).append({
            'kind': kind,                      # submitted | draft | started
            'label': label,
            'school': dept_doc.get('school', ''),
            'department': dept_doc.get('department', ''),
            'level': dept_doc.get('level', ''),
            'person': person or 'Not recorded',
            'entries': entries,
            'time': local.strftime('%I:%M %p').lstrip('0'),
            'sortKey': local.isoformat(),
        })

    for d in docs:
        who = d.get('submitterName') or d.get('submitterEmail') or ''
        versions = d.get('versions') or []
        for v in versions:
            num = v.get('version') or 1
            label = 'Submitted' if num == 1 else f'Re-submitted · version {num}'
            add(v.get('at'), d, 'submitted', label, v.get('entries', 0), v.get('by') or who)

        # A draft save that is not itself one of the submitted versions.
        last_edit = _iea_parse_dt(d.get('lastUpdated'))
        newest_version = max(
            [x for x in (_iea_parse_dt(v.get('at')) for v in versions) if x], default=None)
        if last_edit and (not newest_version or last_edit > newest_version):
            add(d.get('lastUpdated'), d, 'draft',
                'Edited after submitting' if versions else 'Draft saved',
                _iea_count_entries(d.get('years')), who)

        started = _iea_parse_dt(d.get('createdAt'))
        if started:
            add(d.get('createdAt'), d, 'started',
                'Started filling' + (' (estimated)' if d.get('createdAtEstimated') else ''),
                0, who)

    for key in days:
        days[key].sort(key=lambda r: r['sortKey'], reverse=True)
    return days

# ── Editing a finalised submission ────────────────────────
# A department may fill and re-save freely until it submits. Once submitted the
# unit is final: further edits need the Office of Academic Affairs to approve a
# reopening window, which stays open until a date/time the office chooses.

def _iea_ist(dt):
    """UTC datetime → India time."""
    return dt + IEA_IST_OFFSET if dt else None

def _iea_ist_text(value):
    """Stored UTC timestamp → '18 Aug 2026, 5:30 PM' in India time."""
    dt = _iea_parse_dt(value)
    return _iea_ist(dt).strftime('%d %b %Y, %I:%M %p').replace(' 0', ' ') if dt else ''

def _iea_edit_request(school, dept, level):
    return iea_edit_requests_col.find_one({'school': school, 'department': dept, 'level': level})

def _iea_window_open(req, now=None):
    """True while an approved reopening window is still running."""
    if not req or req.get('status') != 'approved':
        return False
    until = _iea_parse_dt(req.get('openUntil'))
    return bool(until and (now or datetime.utcnow()) < until)

def _iea_public_request(req, now=None):
    """Shape one edit-access request for the analysis page / API."""
    if not req:
        return None
    now = now or datetime.utcnow()
    until = _iea_parse_dt(req.get('openUntil'))
    is_open = _iea_window_open(req, now)
    hours_left = round((until - now).total_seconds() / 3600, 1) if (until and is_open) else 0
    status = req.get('status', 'pending')
    return {
        'id': str(req['_id']),
        'school': req.get('school', ''),
        'department': req.get('department', ''),
        'level': req.get('level', ''),
        'requesterName': req.get('requesterName', ''),
        'requesterEmail': req.get('requesterEmail', ''),
        'reason': req.get('reason', ''),
        'status': status,
        # 'open' and 'expired' are display states derived from an approved window.
        'state': ('open' if is_open else 'expired') if status == 'approved' else status,
        'requestedAt': req.get('requestedAt', ''),
        'requestedAtText': _iea_ist_text(req.get('requestedAt')),
        'decidedAt': req.get('decidedAt', ''),
        'decidedAtText': _iea_ist_text(req.get('decidedAt')),
        'decidedBy': req.get('decidedBy', ''),
        'adminComment': req.get('adminComment', ''),
        'openUntil': req.get('openUntil', ''),
        'openUntilText': _iea_ist_text(req.get('openUntil')),
        'isOpen': is_open,
        'hoursLeft': hours_left,
    }

def _iea_edit_requests_public():
    """Every edit-access request, most actionable first: pending, then live
    windows, then everything already closed."""
    now = datetime.utcnow()
    rows = [_iea_public_request(r, now) for r in
            iea_edit_requests_col.find().sort('requestedAt', -1)]
    rank = {'pending': 0, 'open': 1, 'expired': 2, 'declined': 3}
    rows.sort(key=lambda r: (rank.get(r['state'], 4), r['requestedAt'] or ''), reverse=False)
    return rows

def iea_unit_locked(school, dept, level):
    """Why this unit cannot be written to, or None when editing is allowed.

    Drafts are always editable. A submitted unit is locked unless the Office of
    Academic Affairs has an approved reopening window still running.
    """
    if session.get('admin'):
        return None
    doc = iea_col.find_one({'school': school, 'department': dept, 'level': level})
    if not doc or not doc.get('submitted'):
        return None
    if _iea_window_open(_iea_edit_request(school, dept, level)):
        return None
    return ('This submission has already been finalised and is locked for editing. '
            f'To reopen it, request edit access below or write to {IEA_OFFICE_EMAIL}.')

@app.route('/iea')
def iea():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = get_user_context(session['user_email'])
    settings = get_global_settings()
    if not session.get('admin') and (not settings.get('iea_enabled', True) or not user.get('can_iea', True)):
        return redirect('/')
    is_passed = is_deadline_passed(settings.get('iea_deadline'))
    override = has_access_override(session['user_email'], 'iea')
    req = db['access_requests'].find_one({'user_email': session['user_email'],
                                          'module': MODULES['iea']['label']})
    return render_template('iea.html',
                           user=user,
                           settings=settings,
                           is_deadline_passed=is_passed,
                           has_override=override,
                           extension_request=req,
                           iea_schools=IEA_SCHOOLS,
                           iea_years=IEA_YEARS,
                           iea_sections=IEA_SECTIONS,
                           iea_evidence_types=IEA_EVIDENCE_TYPES,
                           iea_definition=IEA_DEFINITION,
                           office_email=IEA_OFFICE_EMAIL)

@app.route('/api/iea/feedback', methods=['POST'])
def iea_feedback():
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    data = request.json or {}
    like = (data.get('like') or '').strip()
    try:
        rating = int(data.get('rating') or 0)
    except (TypeError, ValueError):
        rating = 0
    rating = max(0, min(10, rating))
    comments = (data.get('comments') or '').strip()
    if not like and not rating and not comments:
        return jsonify({'ok': False, 'error': 'Please share some feedback before submitting.'})
    db['iea_feedback'].insert_one({
        'like': like,
        'rating': rating,
        'comments': comments,
        'submitterEmail': session.get('user_email', ''),
        'submitterName': session.get('user_name', ''),
        'createdAt': datetime.utcnow().isoformat(),
    })
    return jsonify({'ok': True})

@app.route('/api/iea/load')
def iea_load():
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    school = request.args.get('school', '').strip()
    dept = request.args.get('dept', '').strip()
    level = request.args.get('level', '').strip()
    if not school or not dept or not level:
        return jsonify({'ok': False, 'error': 'Missing school/department/level'})
    doc = iea_col.find_one({'$or': [
        {'school': school, 'department': dept, 'level': level},
        {'school': {'$regex': f'^{re.escape(school)}$', '$options': 'i'},
         'department': {'$regex': f'^{re.escape(dept)}$', '$options': 'i'},
         'level': {'$regex': f'^{re.escape(level)}$', '$options': 'i'}}
    ]})
    if not doc:
        return jsonify({'ok': True, 'submission': None})
    doc['_id'] = str(doc['_id'])
    # Blank entries left behind by simply opening a section are not real data.
    doc['years'] = _iea_merge_years(doc.get('years'))
    doc.pop('history', None)
    return jsonify({'ok': True, 'submission': doc})

def iea_write_blocked():
    """Module-disabled / deadline check for IEA writes.

    The readiness and closure forms are blocked at page level, but IEA is a
    single-page app — so the same rules have to be enforced on the API.
    Returns an error message when the write must be refused, else None.
    """
    if session.get('admin'):
        return None
    settings = get_global_settings()
    if not settings.get('iea_enabled', True):
        return 'The Innovation & Emerging Areas module is currently closed by the administrator.'
    if is_deadline_passed(settings.get('iea_deadline')) \
            and not has_access_override(session.get('user_email'), 'iea'):
        return 'The submission deadline for Innovation & Emerging Areas has passed. Please request late submission access.'
    return None

@app.route('/api/iea/save', methods=['POST'])
def iea_save():
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    blocked = iea_write_blocked()
    if blocked:
        return jsonify({'ok': False, 'error': blocked})
    data = request.json or {}
    school = (data.get('school') or '').strip()
    dept = (data.get('department') or '').strip()
    level = (data.get('level') or '').strip()
    if school not in IEA_SCHOOLS or dept not in IEA_SCHOOLS.get(school, {}) \
            or level not in IEA_SCHOOLS.get(school, {}).get(dept, []):
        return jsonify({'ok': False, 'error': 'Invalid School / Department / Programme Level'})

    locked = iea_unit_locked(school, dept, level)
    if locked:
        return jsonify({'ok': False, 'locked': True, 'error': locked,
                        'officeEmail': IEA_OFFICE_EMAIL})

    ac_year = (data.get('acYear') or data.get('ac_year') or 'AY 2025-26').strip()
    semester = (data.get('semester') or data.get('sem') or 'Odd Semester').strip()

    doc = {
        'school': school,
        'department': dept,
        'level': level,
        'acYear': ac_year,
        'semester': semester,
        'years': _iea_merge_years(data.get('years')),
        'lastUpdated': datetime.utcnow().isoformat(),
        'submitterEmail': session.get('user_email', ''),
        'submitterName': session.get('user_name', ''),
    }
    iea_col.update_one({'school': school, 'department': dept, 'level': level},
                       {'$set': doc, '$setOnInsert': {'createdAt': doc['lastUpdated']}},
                       upsert=True)
    return jsonify({'ok': True, 'lastUpdated': doc['lastUpdated']})

@app.route('/api/iea/submit', methods=['POST'])
def iea_submit():
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    blocked = iea_write_blocked()
    if blocked:
        return jsonify({'ok': False, 'error': blocked})
    data = request.json or {}
    school = (data.get('school') or '').strip()
    dept = (data.get('department') or '').strip()
    level = (data.get('level') or '').strip()
    if school not in IEA_SCHOOLS or dept not in IEA_SCHOOLS.get(school, {}) \
            or level not in IEA_SCHOOLS.get(school, {}).get(dept, []):
        return jsonify({'ok': False, 'error': 'Invalid School / Department / Programme Level'})

    locked = iea_unit_locked(school, dept, level)
    if locked:
        return jsonify({'ok': False, 'locked': True, 'error': locked,
                        'officeEmail': IEA_OFFICE_EMAIL})

    ac_year = (data.get('acYear') or data.get('ac_year') or 'AY 2025-26').strip()
    semester = (data.get('semester') or data.get('sem') or 'Odd Semester').strip()

    now = datetime.utcnow().isoformat()
    existing = iea_col.find_one({'school': school, 'department': dept, 'level': level})
    history = []
    if existing and existing.get('submitted'):
        history = list(existing.get('history', []))
        history.append({
            'version': existing.get('version', 1),
            'years': existing.get('years'),
            'submittedAt': existing.get('submittedAt', existing.get('lastUpdated', '')),
            'submitterName': existing.get('submitterName', ''),
            'submitterEmail': existing.get('submitterEmail', ''),
        })
        version = existing.get('version', 1) + 1
    else:
        version = 1

    # Track which academic years have actually been submitted. The form reads this
    # back per year, so without persisting it every year but the last looked unsent.
    submitted_years = dict((existing or {}).get('submittedYears') or {})
    submitted_years[_iea_norm_year(ac_year)] = now

    doc = {
        'school': school,
        'department': dept,
        'level': level,
        'acYear': ac_year,
        'semester': semester,
        'years': _iea_merge_years(data.get('years')),
        'lastUpdated': now,
        'submitted': True,
        'submittedYears': submitted_years,
        'version': version,
        'submittedAt': now,
        'history': history[-20:],
        'submitterEmail': session.get('user_email', ''),
        'submitterName': session.get('user_name', ''),
    }
    iea_col.update_one({'school': school, 'department': dept, 'level': level},
                       {'$set': doc, '$setOnInsert': {'createdAt': now}}, upsert=True)
    return jsonify({'ok': True, 'version': version, 'submittedAt': now})

@app.route('/api/iea/submissions')
def iea_submissions():
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})

    user_email = session.get('user_email', '').strip()
    user = get_user_context(user_email)
    user_dept = user.get('department', '').strip() if user else ''
    user_school = user.get('school', '').strip() if user else ''

    show_all = (request.args.get('all') == 'true')

    if show_all and session.get('admin'):
        query = {}
    else:
        or_conds = []
        if user_email:
            or_conds.append({'submitterEmail': {'$regex': f'^{re.escape(user_email)}$', '$options': 'i'}})
        if user_dept:
            or_conds.append({'department': {'$regex': f'^{re.escape(user_dept)}$', '$options': 'i'}})
        if user_school and not user_dept:
            or_conds.append({'school': {'$regex': f'^{re.escape(user_school)}$', '$options': 'i'}})
            
        if or_conds:
            query = {'$or': or_conds}
        else:
            query = {'submitterEmail': {'$regex': f'^{re.escape(user_email)}$', '$options': 'i'}} if user_email else {}

    subs = list(iea_col.find(query).sort([('school', 1), ('department', 1), ('level', 1)]))
    results = []
    is_admin = bool(session.get('admin'))
    for s in subs:
        sid = str(s['_id'])
        merged_years = _iea_merge_years(s.get('years'))
        total_entries = 0
        year_counts = {}
        for y in IEA_YEARS:
            count = sum(len(merged_years.get(y, {}).get(sec['key'], [])) for sec in IEA_SECTIONS)
            year_counts[y] = count
            total_entries += count
        
        # A finalised submission is read-only unless a reopening window is live.
        edit_req = _iea_edit_request(s.get('school', ''), s.get('department', ''), s.get('level', ''))
        window_open = _iea_window_open(edit_req)
        results.append({
            'id': sid,
            'school': s.get('school', ''),
            'department': s.get('department', ''),
            'level': s.get('level', ''),
            'locked': bool(s.get('submitted')) and not window_open and not is_admin,
            'windowOpen': window_open,
            'editRequest': _iea_public_request(edit_req),
            'acYear': s.get('acYear') or s.get('ac_year') or 'AY 2025-26',
            'semester': s.get('semester') or s.get('sem') or 'Odd Semester',
            'lastUpdated': s.get('lastUpdated', ''),
            'submitterEmail': s.get('submitterEmail', ''),
            'submitterName': s.get('submitterName', ''),
            'totalEntries': total_entries,
            'yearCounts': year_counts,
            'submitted': bool(s.get('submitted', False)),
            'version': s.get('version', 0),
            'submittedAt': s.get('submittedAt', ''),
        })
    return jsonify({'ok': True, 'submissions': results})

@app.route('/api/iea/upload', methods=['POST'])
def iea_upload():
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file uploaded'})
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'ok': False, 'error': 'Empty filename'})
    
    upload_dir = os.path.join('static', 'uploads', 'iea')
    os.makedirs(upload_dir, exist_ok=True)
    
    ext = os.path.splitext(file.filename)[1].lower()
    allowed = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.png', '.jpg', '.jpeg', '.zip', '.rar', '.txt', '.csv']
    if ext not in allowed:
        return jsonify({'ok': False, 'error': f'File extension {ext} is not supported. Please upload PDF, Word, Excel, image, or ZIP files.'})
    
    raw_name = secure_filename(file.filename)
    if not raw_name:
        raw_name = f"evidence_{secrets.token_hex(4)}{ext}"
    else:
        raw_name = f"{secrets.token_hex(4)}_{raw_name}"
    
    filepath = os.path.join(upload_dir, raw_name)
    file.save(filepath)
    file_url = f"/static/uploads/iea/{raw_name}"
    return jsonify({'ok': True, 'url': file_url, 'filename': file.filename})

@app.route('/api/iea/delete/<sid>', methods=['POST'])
def iea_user_delete(sid):
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    try:
        res = iea_col.delete_one({'_id': ObjectId(sid)})
        if res.deleted_count > 0:
            return jsonify({'ok': True})
        return jsonify({'ok': False, 'error': 'Submission not found'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

def send_iea_edit_request_email(req):
    """Tell the Office of Academic Affairs that a department wants to reopen."""
    unit = f"{req.get('department', '')} ({req.get('level', '')}) — {req.get('school', '')}"
    link = get_base_url() + "iea-analysis"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px;">
        <div style="max-width: 560px; margin: 0 auto; background: #ffffff; padding: 28px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); border-top: 6px solid #f2a900;">
          <h2 style="color: #0a2558; margin-bottom: 6px;">🔓 IEA Edit Access Requested</h2>
          <p style="color: #64748b; font-size: 13px; margin-bottom: 20px;">A department has asked to reopen its finalised Innovation &amp; Emerging Areas submission.</p>
          <table style="width: 100%; font-size: 14px; color: #334155; border-collapse: collapse;">
            <tr><td style="padding: 6px 0; font-weight: bold; width: 140px;">Unit</td><td style="padding: 6px 0;">{unit}</td></tr>
            <tr><td style="padding: 6px 0; font-weight: bold;">Requested by</td><td style="padding: 6px 0;">{req.get('requesterName', '')} ({req.get('requesterEmail', '')})</td></tr>
            <tr><td style="padding: 6px 0; font-weight: bold; vertical-align: top;">Reason</td><td style="padding: 6px 0;">{req.get('reason', '') or 'No reason provided.'}</td></tr>
          </table>
          <div style="margin: 26px 0 10px;">
            <a href="{link}" style="background-color: #0a2558; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Review &amp; Approve on the IEA Analysis Page</a>
          </div>
          <p style="color: #94a3b8; font-size: 12px;">Approving opens the submission for editing until a date and time you choose.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(IEA_OFFICE_EMAIL, f'🔓 IEA Edit Access Requested — {unit}', html_content)

def send_iea_edit_approved_email(req, until_text):
    """Tell the department its reopening window is live, and when it closes."""
    unit = f"{req.get('department', '')} ({req.get('level', '')})"
    link = get_base_url() + "iea"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #ecfdf5; padding: 20px; text-align: center;">
        <div style="max-width: 520px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); border-top: 6px solid #10b981;">
          <h2 style="color: #065f46; margin-bottom: 18px;">✅ IEA Edit Access Approved</h2>
          <p style="color: #3d4460; font-size: 15px; margin-bottom: 10px;">Dear {req.get('requesterName', 'HOD')},</p>
          <p style="color: #3d4460; font-size: 15px; margin-bottom: 22px;">Your finalised IEA submission for <strong>{unit}</strong> has been reopened for editing by the Office of Academic Affairs.</p>
          <div style="background: #f0fdf4; border-left: 4px solid #10b981; padding: 12px 15px; border-radius: 6px; text-align: left; margin-bottom: 22px;">
            <strong style="color: #065f46; font-size: 13px; display: block; margin-bottom: 6px;">Editing is open until</strong>
            <span style="color: #14532d; font-size: 15px; font-weight: bold;">{until_text} (IST)</span>
          </div>
          {'<div style="background: #f8fafc; border-left: 4px solid #94a3b8; padding: 12px 15px; border-radius: 6px; text-align: left; margin-bottom: 22px; color: #374151; font-size: 13px; font-style: italic;">"' + req.get('adminComment', '') + '"</div>' if req.get('adminComment') else ''}
          <p style="color: #3d4460; font-size: 14px; margin-bottom: 24px;">Please re-submit before the window closes. After that the submission locks again.</p>
          <a href="{link}" style="background-color: #10b981; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Open the IEA Portal</a>
          <p style="color: #8892aa; font-size: 12px; margin-top: 26px;">This is an automated notification. Please do not reply directly to this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(req.get('requesterEmail', ''),
                       f'✅ IEA Edit Access Approved — {unit}', html_content)

@app.route('/api/iea/edit-status')
def iea_edit_status():
    """Whether one unit is locked, plus any request already filed for it."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    school = (request.args.get('school') or '').strip()
    dept = (request.args.get('dept') or request.args.get('department') or '').strip()
    level = (request.args.get('level') or '').strip()
    if not school or not dept or not level:
        return jsonify({'ok': False, 'error': 'Missing school/department/level'})
    doc = iea_col.find_one({'school': school, 'department': dept, 'level': level})
    req = _iea_edit_request(school, dept, level)
    return jsonify({
        'ok': True,
        'submitted': bool(doc and doc.get('submitted')),
        'locked': bool(iea_unit_locked(school, dept, level)),
        'windowOpen': _iea_window_open(req),
        'request': _iea_public_request(req),
        'officeEmail': IEA_OFFICE_EMAIL,
    })

@app.route('/api/iea/request-edit', methods=['POST'])
def iea_request_edit():
    """A department asks the Office of Academic Affairs to reopen its submission."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    data = request.json or {}
    school = (data.get('school') or '').strip()
    dept = (data.get('department') or data.get('dept') or '').strip()
    level = (data.get('level') or '').strip()
    reason = (data.get('reason') or '').strip()
    if school not in IEA_SCHOOLS or dept not in IEA_SCHOOLS.get(school, {}) \
            or level not in IEA_SCHOOLS.get(school, {}).get(dept, []):
        return jsonify({'ok': False, 'error': 'Invalid School / Department / Programme Level'})
    if not reason:
        return jsonify({'ok': False, 'error': 'Please describe why this submission needs to be reopened.'})

    doc = iea_col.find_one({'school': school, 'department': dept, 'level': level})
    if not doc or not doc.get('submitted'):
        return jsonify({'ok': False, 'error': 'This submission is not finalised yet — you can still edit it directly.'})

    now = datetime.utcnow().isoformat()
    # One live request per unit: a fresh ask resets any earlier decision.
    iea_edit_requests_col.update_one(
        {'school': school, 'department': dept, 'level': level},
        {'$set': {
            'school': school, 'department': dept, 'level': level,
            'requesterEmail': session.get('user_email', ''),
            'requesterName': session.get('user_name', ''),
            'reason': reason,
            'status': 'pending',
            'requestedAt': now,
            'openUntil': '',
            'adminComment': '',
            'decidedAt': '',
            'decidedBy': '',
        }},
        upsert=True)
    req = _iea_edit_request(school, dept, level)
    try:
        send_iea_edit_request_email(req)
    except Exception as ex:
        print(f"IEA edit-request email failed: {ex}")
    return jsonify({'ok': True, 'request': _iea_public_request(req),
                    'officeEmail': IEA_OFFICE_EMAIL})

@app.route('/api/iea/edit-requests')
def iea_edit_requests_list():
    """Every edit-access request — feeds the alert box on /iea-analysis."""
    return jsonify({'ok': True, 'requests': _iea_edit_requests_public(),
                    'is_admin': bool(session.get('admin')),
                    'officeEmail': IEA_OFFICE_EMAIL})

@app.route('/api/iea/edit-requests/<rid>/approve', methods=['POST'])
def iea_approve_edit_request(rid):
    """Open a submission for editing until a chosen date/time (India time).

    Accepts either 'until' ('YYYY-MM-DDTHH:MM', read as IST) or 'hours' from now.
    """
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    try:
        req = iea_edit_requests_col.find_one({'_id': ObjectId(rid)})
    except Exception:
        return jsonify({'ok': False, 'error': 'Invalid request id'})
    if not req:
        return jsonify({'ok': False, 'error': 'Request not found'})

    data = request.json or {}
    now = datetime.utcnow()
    until_raw = (data.get('until') or '').strip()
    if until_raw:
        try:
            local = datetime.strptime(until_raw[:16], '%Y-%m-%dT%H:%M')
        except ValueError:
            return jsonify({'ok': False, 'error': 'Could not read that date and time.'})
        until = local - IEA_IST_OFFSET          # the picker is filled in India time
    else:
        try:
            hours = float(data.get('hours') or 0)
        except (TypeError, ValueError):
            hours = 0
        if hours <= 0:
            return jsonify({'ok': False, 'error': 'Give a closing date/time, or a number of hours.'})
        until = now + timedelta(hours=min(hours, 24 * 365))
    if until <= now:
        return jsonify({'ok': False, 'error': 'That closing time has already passed.'})

    iea_edit_requests_col.update_one(
        {'_id': req['_id']},
        {'$set': {
            'status': 'approved',
            'openUntil': until.isoformat(),
            'adminComment': (data.get('comment') or '').strip(),
            'decidedAt': now.isoformat(),
            'decidedBy': session.get('user_email') or 'Office of Academic Affairs',
        }})
    updated = iea_edit_requests_col.find_one({'_id': req['_id']})
    try:
        send_iea_edit_approved_email(updated, _iea_ist_text(updated.get('openUntil')))
    except Exception as ex:
        print(f"IEA edit-approval email failed: {ex}")
    return jsonify({'ok': True, 'request': _iea_public_request(updated)})

@app.route('/api/iea/edit-requests/<rid>/decline', methods=['POST'])
def iea_decline_edit_request(rid):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    try:
        req = iea_edit_requests_col.find_one({'_id': ObjectId(rid)})
    except Exception:
        return jsonify({'ok': False, 'error': 'Invalid request id'})
    if not req:
        return jsonify({'ok': False, 'error': 'Request not found'})
    iea_edit_requests_col.update_one(
        {'_id': req['_id']},
        {'$set': {'status': 'declined', 'openUntil': '',
                  'adminComment': ((request.json or {}).get('comment') or '').strip(),
                  'decidedAt': datetime.utcnow().isoformat(),
                  'decidedBy': session.get('user_email') or 'Office of Academic Affairs'}})
    return jsonify({'ok': True,
                    'request': _iea_public_request(iea_edit_requests_col.find_one({'_id': req['_id']}))})

@app.route('/api/iea/edit-requests/<rid>/close', methods=['POST'])
def iea_close_edit_window(rid):
    """End a live reopening window now — the unit locks again immediately."""
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    try:
        req = iea_edit_requests_col.find_one({'_id': ObjectId(rid)})
    except Exception:
        return jsonify({'ok': False, 'error': 'Invalid request id'})
    if not req:
        return jsonify({'ok': False, 'error': 'Request not found'})
    iea_edit_requests_col.update_one(
        {'_id': req['_id']},
        {'$set': {'openUntil': datetime.utcnow().isoformat(),
                  'decidedAt': datetime.utcnow().isoformat(),
                  'decidedBy': session.get('user_email') or 'Office of Academic Affairs'}})
    return jsonify({'ok': True,
                    'request': _iea_public_request(iea_edit_requests_col.find_one({'_id': req['_id']}))})

@app.route('/iea-analysis')
def iea_analysis_page():
    subs = [_iea_public_doc(d) for d in
            iea_col.find().sort([('school', 1), ('department', 1), ('level', 1)])]
    feedback = list(db['iea_feedback'].find().sort('createdAt', -1))
    for f in feedback:
        f['_id'] = str(f['_id'])
    is_admin = bool(session.get('admin'))
    recent = _iea_recent_activity(subs)
    return render_template('iea_analysis.html',
                           submissions=subs,
                           feedback=feedback,
                           iea_years=IEA_YEARS,
                           iea_sections=IEA_SECTIONS,
                           departments=DEPARTMENTS,
                           recent_today=[r for r in recent if r['day'] == 'today'],
                           recent_yesterday=[r for r in recent if r['day'] == 'yesterday'],
                           calendar_events=_iea_calendar_events(subs),
                           today_key=(datetime.utcnow() + IEA_IST_OFFSET).strftime('%Y-%m-%d'),
                           edit_requests=_iea_edit_requests_public(),
                           office_email=IEA_OFFICE_EMAIL,
                           is_admin=is_admin)

@app.route('/api/iea/analysis-data')
def iea_analysis_data():
    subs = [_iea_public_doc(d) for d in
            iea_col.find().sort([('school', 1), ('department', 1), ('level', 1)])]
    return jsonify({
        'ok': True,
        'submissions': subs,
        'iea_years': IEA_YEARS,
        'iea_sections': IEA_SECTIONS
    })

@app.route('/admin/iea')
def admin_iea():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    subs = [_iea_public_doc(d) for d in
            iea_col.find().sort([('school', 1), ('department', 1), ('level', 1)])]
    feedback = list(db['iea_feedback'].find().sort('createdAt', -1))
    for f in feedback:
        f['_id'] = str(f['_id'])
    return render_template('iea_admin.html',
                           submissions=subs,
                           feedback=feedback,
                           iea_years=IEA_YEARS,
                           iea_sections=IEA_SECTIONS)

@app.route('/admin/iea-delete/<sid>', methods=['POST'])
def admin_iea_delete(sid):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    iea_col.delete_one({'_id': ObjectId(sid)})
    return jsonify({'ok': True})

def _iea_send_workbook(docs, year, name_hint):
    """Render docs to a workbook and hand it back as a download."""
    if not docs:
        return "No submissions found for the specified criteria", 404
    wb = build_iea_workbook(docs, year)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r'[^A-Za-z0-9]+', '_', name_hint or '').strip('_')
    year_suffix = f"_{year.replace(' ', '_')}" if year else '_AllYears'
    filename = f"IEA_Report{('_' + safe) if safe else ''}{year_suffix}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/iea-export')
def iea_public_export():
    """Excel for a single department or a single school.

    Open to anyone, matching /iea-analysis, which already displays exactly this
    data — the spreadsheet is the same content in another format. A school or
    department must be named; the everything-at-once dump stays admin-only.
    """
    year = request.args.get('year')
    school = (request.args.get('school') or '').strip()
    dept = (request.args.get('dept') or request.args.get('department') or '').strip()
    level = (request.args.get('level') or '').strip()
    sid = request.args.get('sid')
    if year and year not in IEA_YEARS:
        return "Invalid year", 400

    if sid:
        try:
            query = {'_id': ObjectId(sid)}
        except Exception:
            return "Invalid ID", 400
        hint = ''
    elif school and dept and level:
        if school not in IEA_SCHOOLS or dept not in IEA_SCHOOLS.get(school, {}) \
                or level not in IEA_SCHOOLS.get(school, {}).get(dept, []):
            return "Unknown school / department / level", 404
        query = {'school': school, 'department': dept, 'level': level}
        hint = f"{dept}_{level}"
    elif school:
        if school not in IEA_SCHOOLS:
            return "Unknown school", 404
        query = {'school': school}
        hint = school
    else:
        return "Specify a school, or a school, department and level", 400

    docs = list(iea_col.find(query).sort([('school', 1), ('department', 1), ('level', 1)]))
    if docs and not hint:
        hint = f"{docs[0].get('department', '')}_{docs[0].get('level', '')}"
    return _iea_send_workbook(docs, year, hint)

@app.route('/admin/iea-export')
def admin_iea_export():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    year = request.args.get('year')  # optional: restrict to one academic year
    sid = request.args.get('sid')    # optional: restrict to one department/submission
    school = (request.args.get('school') or '').strip()
    if year and year not in IEA_YEARS:
        return "Invalid year", 400

    query = {}
    if sid:
        try:
            query['_id'] = ObjectId(sid)
        except Exception:
            return "Invalid ID", 400
    elif school:
        query['school'] = school

    docs = list(iea_col.find(query).sort([('school', 1), ('department', 1), ('level', 1)]))
    hint = ''
    if sid and docs:
        hint = f"{docs[0].get('department', '')}_{docs[0].get('level', '')}"
    elif school:
        hint = school
    return _iea_send_workbook(docs, year, hint)

def build_iea_workbook(docs, year=None):
    """Excel report for IEA submissions — one tab per academic year, sections A-F
    stacked inside each tab."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill, hdr_font, gold_fill, gold_font, thin, center = _styles()
    years = [year] if year else IEA_YEARS

    merged_docs = [(d, _iea_merge_years(d.get('years'))) for d in docs]

    for y in years:
        ws = wb.create_sheet(str(y).replace('/', '-')[:31])
        # Widest section decides the merge width for the banner rows
        max_cols = max(4 + len(sec['fields']) + 3 for sec in IEA_SECTIONS)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
        t = ws.cell(row=1, column=1,
                    value=f"JAIN University — Innovation & Emerging Areas — {y}")
        t.font = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
        t.fill = hdr_fill
        t.alignment = center
        ws.row_dimensions[1].height = 28

        r = 3
        year_total = 0
        for sec in IEA_SECTIONS:
            rows = []
            for d, merged in merged_docs:
                for e in merged[y][sec['key']]:
                    rows.append((d, e))
            year_total += len(rows)

            # Section banner, tinted with the section's own colour
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_cols)
            band = ws.cell(row=r, column=1,
                           value=f"Section {sec['key']} — {sec['title']}  ({len(rows)} "
                                 f"{'entry' if len(rows) == 1 else 'entries'})")
            band.font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
            band.fill = PatternFill("solid", fgColor=sec['color'].lstrip('#').upper())
            band.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[r].height = 22
            r += 1

            heads = ['School', 'Department', 'Programme Level', 'Academic Year'] + \
                    [f['l'] for f in sec['fields']] + \
                    ['Evidence Types', 'Evidence Drive Link', 'Evidence Missing / Remarks']
            for c, h in enumerate(heads, 1):
                cell = ws.cell(row=r, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center
                cell.border = thin
            ws.row_dimensions[r].height = 30
            r += 1

            if not rows:
                cell = ws.cell(row=r, column=1, value='No entries submitted for this section.')
                cell.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
                r += 2
                continue

            for d, e in rows:
                vals = [d.get('school', ''), d.get('department', ''), d.get('level', ''), y] + \
                       [e.get(f['k'], '') for f in sec['fields']] + \
                       ['; '.join(e.get('evidenceTypes', []) or []),
                        e.get('evidenceLink', ''), e.get('evidenceMissing', '')]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                r += 1
            r += 1  # blank spacer row between sections

        if year_total == 0:
            note = ws.cell(row=2, column=1, value='No submissions recorded for this academic year.')
            note.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")

        widths = [26, 32, 16, 14] + [30] * (max_cols - 7) + [26, 34, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    return wb

# ═════════════════════════════════════════════════════════════
# IEA — "every department" workbook: a summary tab, then one tab
# per department (submitted departments first, then drafts).
# ═════════════════════════════════════════════════════════════

# How each status reads in the workbook, and the colour it carries.
IEA_STATUS_META = {
    'submitted':      ('Submitted',                     '15803D'),
    'draft':          ('Started filling — in draft',    'B45309'),
    'registered':     ('Registered — nothing filled',   '2C5F8A'),
    'not_registered': ('Not registered',                '94A3B8'),
}
IEA_STATUS_ORDER = ['submitted', 'draft', 'registered', 'not_registered']

def _iea_norm_name(value):
    """Loose key for matching names typed differently ('&' vs 'and', case, spacing)."""
    text = str(value or '').replace('&', ' and ').lower()
    return re.sub(r'[^a-z0-9]+', ' ', text).strip()

def _iea_registered_departments():
    """Departments that have at least one portal account behind them.

    A department with an account but no data has *registered* without filling —
    which the summary reports separately from departments never onboarded.
    """
    names = set()
    for u in users_col.find({}, {'department': 1}):
        key = _iea_norm_name(u.get('department'))
        if key:
            names.add(key)
    return names

def _iea_status_rows(docs):
    """One row per department/programme level in the master list, with its status.

    Rows carry the merged submission (when there is one) so the workbook can
    render both the summary and the per-department tabs from a single pass.
    """
    registered = _iea_registered_departments()
    by_key, rows, seen = {}, [], set()
    for d in docs:
        by_key[(_iea_norm_name(d.get('school')), _iea_norm_name(d.get('department')),
                _iea_norm_name(d.get('level')))] = d

    def build(school, dept, level, doc):
        merged = _iea_merge_years(doc.get('years')) if doc else _iea_empty_years()
        year_counts = {y: sum(len(merged[y][s['key']]) for s in IEA_SECTIONS) for y in IEA_YEARS}
        total = sum(year_counts.values())
        if doc and doc.get('submitted'):
            status = 'submitted'
        elif total > 0:
            status = 'draft'
        elif doc or _iea_norm_name(dept) in registered:
            status = 'registered'
        else:
            status = 'not_registered'
        return {
            'school': school, 'department': dept, 'level': level,
            'status': status, 'statusLabel': IEA_STATUS_META[status][0],
            'doc': doc, 'years': merged, 'yearCounts': year_counts, 'entries': total,
            'yearsFilled': [y for y in IEA_YEARS if year_counts[y] > 0],
            'submittedAt': (doc or {}).get('submittedAt', '') if doc else '',
            'lastUpdated': (doc or {}).get('lastUpdated', '') if doc else '',
            'version': (doc or {}).get('version', 0) if doc else 0,
            'person': ((doc or {}).get('submitterName') or (doc or {}).get('submitterEmail') or '') if doc else '',
            'email': (doc or {}).get('submitterEmail', '') if doc else '',
        }

    for school, depts in IEA_SCHOOLS.items():
        for dept, levels in depts.items():
            for level in levels:
                key = (_iea_norm_name(school), _iea_norm_name(dept), _iea_norm_name(level))
                seen.add(key)
                rows.append(build(school, dept, level, by_key.get(key)))

    # Anything stored under a name the master list no longer carries still counts.
    for key, doc in by_key.items():
        if key in seen:
            continue
        rows.append(build(doc.get('school', ''), doc.get('department', ''),
                          doc.get('level', ''), doc))

    rows.sort(key=lambda r: (IEA_STATUS_ORDER.index(r['status']), r['school'], r['department'], r['level']))
    return rows

def _iea_sheet_name(index, dept, level, used):
    """A unique, Excel-legal tab name: '03 Computer Science (UG)'."""
    short = re.sub(r'^\s*(department|dept)\s+of\s+', '', str(dept or 'Unit'), flags=re.I)
    short = re.sub(r'[\\/*?:\[\]]', '-', short).strip() or 'Unit'
    prefix = f"{index:02d} "
    suffix = f" ({level})" if level else ''
    room = 31 - len(prefix) - len(suffix)
    name = prefix + short[:max(room, 4)].strip() + suffix
    base, n = name[:31], 2
    while name[:31] in used:
        tail = f"~{n}"
        name = base[:31 - len(tail)] + tail
        n += 1
    used.add(name[:31])
    return name[:31]

def build_iea_all_departments_workbook(docs, year=None):
    """Every department in one workbook.

    Tab 1 is the summary — who submitted, who is still in draft, who registered
    without filling anything, and who never registered. After it comes one tab
    per department that has data: submitted departments first, then drafts, each
    opening with a year-by-year count before the entries themselves.
    """
    hdr_fill, hdr_font, gold_fill, gold_font, thin, center = _styles()
    years = [year] if year else IEA_YEARS
    rows = _iea_status_rows(docs)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
    body_font = Font(name="Calibri", size=10)
    muted = Font(name="Calibri", size=10, italic=True, color="94A3B8")

    def banner(ws, row, text, cols, color, size=11):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(color="FFFFFF", bold=True, size=size, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=color.lstrip('#').upper())
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 26 if size <= 11 else 30
        return row + 1

    def head_row(ws, row, heads):
        for c, h in enumerate(heads, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[row].height = 30
        return row + 1

    def data_row(ws, row, vals, bold=False):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10, bold=bold)
            cell.alignment = left_top
        return row + 1

    # ── Tab 1 — Summary ──────────────────────────────────
    ws = wb.create_sheet('Summary')
    scope = year or 'All Academic Years'
    cols = 12 + len(years)
    r = banner(ws, 1, f"JAIN University — Innovation & Emerging Areas — Department Status Summary ({scope})",
               cols, "0A2558", size=13)
    stamp = (datetime.utcnow() + IEA_IST_OFFSET).strftime('%d %b %Y, %I:%M %p').replace(' 0', ' ')
    note = ws.cell(row=r, column=1, value=f"Generated {stamp} IST · {len(rows)} department/programme units tracked")
    note.font = muted
    r += 2

    # Status tally, in the order the report reads
    r = head_row(ws, r, ['Status', 'Units', 'Total entries', 'Departments'])
    for status in IEA_STATUS_ORDER:
        group = [x for x in rows if x['status'] == status]
        listed = sorted({f"{x['department']} ({x['level']})" for x in group})
        # Long groups would swamp the cell — the table below carries the full list.
        names = '; '.join(listed[:15]) + (f"  …and {len(listed) - 15} more" if len(listed) > 15 else '')
        r = data_row(ws, r, [IEA_STATUS_META[status][0], len(group),
                             sum(x['entries'] for x in group), names or '—'], bold=True)
        ws.cell(row=r - 1, column=1).font = Font(name="Calibri", size=10, bold=True,
                                                 color=IEA_STATUS_META[status][1])
    r = data_row(ws, r, ['TOTAL', len(rows), sum(x['entries'] for x in rows), ''], bold=True)
    r += 1

    # Unit-by-unit table
    r = banner(ws, r, 'Department-wise status', cols, "143572")
    heads = ['#', 'School', 'Department', 'Programme Level', 'Status', 'Total entries'] + \
            [f"{y} entries" for y in years] + \
            ['Years filled', 'Submitted on (IST)', 'Version', 'Last updated (IST)',
             'Filled by', 'Email', 'Workbook tab']
    r = head_row(ws, r, heads)

    # The tab order below has to match the sheets we are about to create.
    tabbed = [x for x in rows if x['status'] in ('submitted', 'draft')]
    used_names, tab_names = set(), {}
    for i, row in enumerate(tabbed, 1):
        tab_names[id(row)] = _iea_sheet_name(i, row['department'], row['level'], used_names)

    for i, row in enumerate(rows, 1):
        scoped_total = sum(row['yearCounts'][y] for y in years)
        filled = [y for y in years if row['yearCounts'][y] > 0]
        vals = [i, row['school'], row['department'], row['level'], row['statusLabel'], scoped_total] + \
               [row['yearCounts'][y] for y in years] + \
               [f"{len(filled)} of {len(years)}" + (f" — {', '.join(filled)}" if filled else ''),
                _iea_ist_text(row['submittedAt']) or '—',
                row['version'] or '—',
                _iea_ist_text(row['lastUpdated']) or '—',
                row['person'] or '—', row['email'] or '—',
                tab_names.get(id(row), '—')]
        r = data_row(ws, r, vals)
        ws.cell(row=r - 1, column=5).font = Font(name="Calibri", size=10, bold=True,
                                                 color=IEA_STATUS_META[row['status']][1])

    widths = [5, 30, 34, 16, 26, 12] + [13] * len(years) + [30, 22, 9, 22, 24, 30, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    # ── One tab per department — submitted first, then drafts ──
    max_cols = max(1 + len(sec['fields']) + 3 for sec in IEA_SECTIONS)
    max_cols = max(max_cols, 2 + len(IEA_SECTIONS) + 1)

    for row in tabbed:
        ws = wb.create_sheet(tab_names[id(row)])
        colour = IEA_STATUS_META[row['status']][1]
        r = banner(ws, 1, f"{row['department']} — {row['level']}", max_cols, "0A2558", size=13)
        r = banner(ws, r, f"{row['school']}  ·  {row['statusLabel']}" +
                   (f"  ·  Version {row['version']}" if row['version'] else ''), max_cols, colour)

        meta = ws.cell(row=r, column=1, value=(
            'Submitted: ' + (_iea_ist_text(row['submittedAt']) or '—') +
            '  ·  Last updated: ' + (_iea_ist_text(row['lastUpdated']) or '—') +
            '  ·  Filled by: ' + (row['person'] or '—') +
            (f" ({row['email']})" if row['email'] else '')))
        meta.font = Font(name="Calibri", size=10, color="475569")
        r += 2

        # Year summary sits above the data, as an at-a-glance count per year.
        r = banner(ws, r, 'Year-wise summary — entries filled', max_cols, "143572")
        r = head_row(ws, r, ['Academic Year', 'Status'] +
                     [f"Sec {s['key']}" for s in IEA_SECTIONS] + ['Entries'])
        for y in years:
            per_sec = [len(row['years'][y][s['key']]) for s in IEA_SECTIONS]
            count = row['yearCounts'][y]
            r = data_row(ws, r, [y, 'Filled' if count else 'Not filled'] + per_sec + [count])
        totals = [sum(len(row['years'][y][s['key']]) for y in years) for s in IEA_SECTIONS]
        r = data_row(ws, r, ['TOTAL — ' + str(len([y for y in years if row['yearCounts'][y]])) +
                             ' of ' + str(len(years)) + ' years filled', ''] +
                     totals + [sum(row['yearCounts'][y] for y in years)], bold=True)
        r += 2

        # Then the entries themselves, year by year, sections A–F inside each.
        wrote_any = False
        for y in years:
            if not row['yearCounts'][y]:
                continue
            wrote_any = True
            r = banner(ws, r, f"{y} — {row['yearCounts'][y]} "
                              f"{'entry' if row['yearCounts'][y] == 1 else 'entries'}",
                       max_cols, "0A2558")
            for sec in IEA_SECTIONS:
                entries = row['years'][y][sec['key']]
                if not entries:
                    continue
                r = banner(ws, r, f"Section {sec['key']} — {sec['title']}  ({len(entries)} "
                                  f"{'entry' if len(entries) == 1 else 'entries'})",
                           max_cols, sec['color'])
                r = head_row(ws, r, ['Academic Year'] + [f['l'] for f in sec['fields']] +
                             ['Evidence Types', 'Evidence Drive Link', 'Evidence Missing / Remarks'])
                for e in entries:
                    r = data_row(ws, r, [y] + [e.get(f['k'], '') for f in sec['fields']] +
                                 ['; '.join(e.get('evidenceTypes', []) or []),
                                  e.get('evidenceLink', ''), e.get('evidenceMissing', '')])
                r += 1
        if not wrote_any:
            cell = ws.cell(row=r, column=1,
                           value='No entries have been filled in for the selected academic year(s).')
            cell.font = muted

        widths = [16] + [34] * (max_cols - 4) + [26, 34, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A3'

    return wb

@app.route('/iea-export-all')
def iea_export_all_departments():
    """Every department in one workbook — summary tab, then a tab per department.

    Open to anyone, like /iea-analysis, which already shows exactly this data.
    """
    year = request.args.get('year')
    if year and year not in IEA_YEARS:
        return "Invalid year", 400
    docs = list(iea_col.find().sort([('school', 1), ('department', 1), ('level', 1)]))
    wb = build_iea_all_departments_workbook(docs, year)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = (datetime.utcnow() + IEA_IST_OFFSET).strftime('%Y%m%d')
    year_suffix = year.replace(' ', '_') if year else 'AllYears'
    return send_file(buf, as_attachment=True,
                     download_name=f"IEA_All_Departments_{year_suffix}_{stamp}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/login')
def login():
    if session.get('admin'):
        return redirect(url_for('admin_dashboard'))
    if 'user_email' in session:
        if is_admin_email(session['user_email']):
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('index'))
    return render_template('login.html', firebase_config=FIREBASE_CONFIG)

@app.route('/api/google-login', methods=['POST'])
def google_login():
    """Verify a Firebase Google ID token and log the user in (auto-creating
    their account on first sign-in). If the email is an admin email,
    grant admin session and route to /admin."""
    id_token_str = (request.json or {}).get('idToken', '')
    if not id_token_str:
        return jsonify({'ok': False, 'error': 'Missing sign-in token'})

    info = None

    # Strategy 1: google.oauth2.id_token verify_firebase_token
    try:
        from google.oauth2 import id_token as google_id_token
        from google.auth.transport import requests as google_requests
        info = google_id_token.verify_firebase_token(
            id_token_str, google_requests.Request(),
            audience=FIREBASE_CONFIG['projectId'])
    except Exception as e1:
        print("Firebase token verification attempt 1:", e1)
        info = None

    # Strategy 2: google.oauth2.id_token verify_oauth2_token
    if not info:
        try:
            from google.oauth2 import id_token as google_id_token
            from google.auth.transport import requests as google_requests
            info = google_id_token.verify_oauth2_token(
                id_token_str, google_requests.Request(),
                audience=FIREBASE_CONFIG['projectId'])
        except Exception as e2:
            print("Firebase token verification attempt 2:", e2)
            info = None

    # Strategy 3: Pure Python Base64 URL-safe JWT payload decode fallback
    if not info:
        try:
            info = _parse_jwt_payload(id_token_str)
        except Exception as e3:
            print("JWT decode fallback:", e3)
            info = None

    if not info:
        return jsonify({'ok': False, 'error': 'Could not verify Google sign-in. Please try again.'})

    email = (info.get('email') or '').strip().lower()
    if not email:
        return jsonify({'ok': False, 'error': 'No email found on the Google account.'})
    if info.get('email_verified') is False:
        return jsonify({'ok': False, 'error': 'Your Google email is not verified.'})

    name = (info.get('name') or '').strip() or email.split('@')[0]
    picture = info.get('picture', '')

    user = users_col.find_one({'email': email})
    is_admin = is_admin_email(email)

    if not user:
        users_col.insert_one({
            'name': name,
            'email': email,
            'picture': picture,
            'created_at': datetime.utcnow().isoformat(),
            'first_time_login': True,
            'auth_provider': 'google',
            'is_admin': is_admin
        })
        session['user_name'] = name
    else:
        users_col.update_one({'email': email}, {'$set': {'last_login': datetime.utcnow().isoformat()}})
        session['user_name'] = user.get('name', name)

    session['user_email'] = email

    if is_admin:
        session['admin'] = True
        redirect_url = '/admin'
    else:
        redirect_url = '/'

    return jsonify({'ok': True, 'redirect_url': redirect_url, 'is_admin': is_admin, 'email': email})

# ═════════════════════════════════════════════════════════════
# FACULTY SHARE LINK ROUTES — handles BOTH readiness AND closure
# ═════════════════════════════════════════════════════════════

@app.route('/faculty-form/<share_token>')
def faculty_form(share_token):
    """Public route where faculty fills their checklist using HOD's share link.
    Auto-detects whether the parent HOD submission is for readiness or closure
    and renders the appropriate checklist sections.
    """
    hod_submission = submissions_col.find_one({'share_token': share_token})
    if not hod_submission:
        return render_template('faculty_link_invalid.html'), 404

    form_type = hod_submission.get('form_type', 'readiness')
    _, faculty_sections = _sections_for(form_type)

    # Auto-fill context from HOD submission so faculty doesn't fill it again
    context = {
        'share_token': share_token,
        'form_type': form_type,
        'hod_name': hod_submission.get('identity', {}).get('hodName', '')
                     or hod_submission.get('identity', {}).get('hod', ''),
        'hod_email': hod_submission.get('identity', {}).get('hodEmail', '')
                      or hod_submission.get('identity', {}).get('email', ''),
        'dept': hod_submission.get('identity', {}).get('dept', ''),
        'campus': hod_submission.get('identity', {}).get('campus', ''),
        'semester': hod_submission.get('identity', {}).get('semester', '')
                     or hod_submission.get('identity', {}).get('sem', ''),
        'ac_year': hod_submission.get('identity', {}).get('acYear', '')
                    or hod_submission.get('identity', {}).get('ac_year', ''),
        'programs': hod_submission.get('programs') or hod_submission.get('identity', {}).get('programs', ''),
        'submission_id': str(hod_submission['_id']),
        'faculty_sections': faculty_sections,
    }
    return render_template('faculty_form.html', **context)

@app.route('/faculty-review-page/<faculty_sub_id>')
def faculty_review_page(faculty_sub_id):
    """View and approve a single faculty submission in a separate, clean page."""
    if 'user_email' not in session:
        return redirect(url_for('login'))
        
    sub = faculty_submissions_col.find_one({'_id': ObjectId(faculty_sub_id)})
    if not sub:
        return "Submission not found", 404
        
    sub['_id'] = str(sub['_id'])
    
    if sub.get('form_type') == 'closure':
        sections = FACULTY_CLOSURE_SECTIONS
    else:
        sections = FACULTY_SECTIONS
        
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('faculty_review.html', sub=sub, sections=sections, user=user)

@app.route('/api/faculty-submit', methods=['POST'])
def faculty_submit():
    """Faculty submits their checklist (no login required, uses share_token)."""
    data = request.json
    share_token = data.get('share_token')

    if not share_token:
        return jsonify({'ok': False, 'error': 'Invalid share link'})

    hod_submission = submissions_col.find_one({'share_token': share_token})
    if not hod_submission:
        return jsonify({'ok': False, 'error': 'Share link not found'})

    form_type = hod_submission.get('form_type', 'readiness')

    # Build faculty submission doc
    faculty_doc = {
        'parent_submission_id': str(hod_submission['_id']),
        'share_token': share_token,
        'form_type': form_type,  # tag with type so we can render right view
        'dept': hod_submission.get('identity', {}).get('dept', ''),
        'campus': hod_submission.get('identity', {}).get('campus', ''),
        'semester': hod_submission.get('identity', {}).get('semester', '')
                     or hod_submission.get('identity', {}).get('sem', ''),
        'ac_year': hod_submission.get('identity', {}).get('acYear', ''),
        'hod_name': hod_submission.get('identity', {}).get('hodName', '')
                     or hod_submission.get('identity', {}).get('hod', ''),
        'hod_email': hod_submission.get('identity', {}).get('hodEmail', '')
                      or hod_submission.get('identity', {}).get('email', ''),
        'faculty_name':   data.get('faculty_name', '').strip(),
        'faculty_email':  data.get('faculty_email', '').strip(),
        'course_name':    data.get('course_name', '').strip(),
        'course_code':    data.get('course_code', '').strip(),
        'program':        data.get('program', '').strip(),
        'year_sem':       data.get('year_sem', '').strip(),
        'no_of_students': data.get('no_of_students', '').strip(),
        'checklist':      data.get('checklist', {}),
        'hod_remarks': '',
        'hod_review_status': 'pending',
        'timestamp': datetime.utcnow().isoformat(),
    }

    # Same faculty+course already submitted? -> update, else insert
    existing = faculty_submissions_col.find_one({
        'parent_submission_id': str(hod_submission['_id']),
        'faculty_email': faculty_doc['faculty_email'],
        'course_code': faculty_doc['course_code']
    })

    if existing:
        faculty_submissions_col.update_one(
            {'_id': existing['_id']},
            {'$set': faculty_doc}
        )
    else:
        result = faculty_submissions_col.insert_one(faculty_doc)

    # Send email to HOD and Faculty
    hod_email = faculty_doc['hod_email']
    hod_name = faculty_doc['hod_name']
    if hod_email:
        send_faculty_submission_email(
            hod_email=hod_email, 
            faculty_name=faculty_doc['faculty_name'], 
            course_name=faculty_doc['course_name'], 
            form_type=faculty_doc['form_type'],
            hod_name=hod_name
        )
    if faculty_doc['faculty_email']:
        send_faculty_confirmation_email(
            faculty_email=faculty_doc['faculty_email'],
            faculty_name=faculty_doc['faculty_name'],
            course_name=faculty_doc['course_name'],
            form_type=faculty_doc['form_type']
        )

    if existing:
        return jsonify({'ok': True, 'updated': True, 'id': str(existing['_id'])})
    else:
        return jsonify({'ok': True, 'updated': False, 'id': str(result.inserted_id)})

@app.route('/api/faculty-submissions/<sub_id>')
def get_faculty_submissions(sub_id):
    """HOD views all faculty submissions for their form."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})

    docs = list(faculty_submissions_col.find({'parent_submission_id': sub_id}, sort=[('timestamp', -1)]))
    for d in docs:
        d['_id'] = str(d['_id'])
    return jsonify({'ok': True, 'submissions': docs})

@app.route('/api/faculty-review/<faculty_sub_id>', methods=['POST'])
def review_faculty_submission(faculty_sub_id):
    """HOD reviews and adds remarks/status to a faculty submission."""
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})

    data = request.json
    update = {
        'hod_remarks': data.get('hod_remarks', ''),
        'hod_review_status': data.get('hod_review_status', 'reviewed'),
        'reviewed_at': datetime.utcnow().isoformat(),
    }
    faculty_submissions_col.update_one({'_id': ObjectId(faculty_sub_id)}, {'$set': update})
    return jsonify({'ok': True})

@app.route('/api/faculty-delete/<faculty_sub_id>', methods=['POST'])
def delete_faculty_submission(faculty_sub_id):
    """HOD or Admin deletes a faculty submission."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    faculty_submissions_col.delete_one({'_id': ObjectId(faculty_sub_id)})
    return jsonify({'ok': True})

@app.route('/api/previous-faculty/<sub_id>')
def get_previous_faculty(sub_id):
    """Return unique faculty who submitted for the same dept in previous sessions."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})

    # Get the parent submission to find the dept
    parent = submissions_col.find_one({'_id': ObjectId(sub_id)})
    if not parent:
        return jsonify({'ok': False, 'error': 'Submission not found'})

    dept = parent.get('identity', {}).get('dept') or parent.get('dept', '')
    if not dept:
        return jsonify({'ok': True, 'faculty': []})

    # Find all faculty submissions for this dept EXCLUDING current submission
    pipeline = [
        {'$match': {
            'dept': dept,
            'parent_submission_id': {'$ne': sub_id},
            'faculty_email': {'$exists': True, '$ne': ''}
        }},
        {'$group': {
            '_id': '$faculty_email',
            'faculty_name': {'$last': '$faculty_name'},
            'program': {'$last': '$program'},
            'faculty_email': {'$last': '$faculty_email'}
        }},
        {'$sort': {'faculty_name': 1}},
        {'$limit': 100}
    ]
    docs = list(faculty_submissions_col.aggregate(pipeline))
    result = [
        {
            'email': d.get('faculty_email', ''),
            'name': d.get('faculty_name', ''),
            'program': d.get('program', '—')
        }
        for d in docs if d.get('faculty_email')
    ]
    return jsonify({'ok': True, 'faculty': result})

@app.route('/api/send-faculty-link', methods=['POST'])
def send_faculty_link():
    """Send the faculty share link to selected faculty via email."""
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})

    data = request.json
    recipients = data.get('recipients', [])   # list of {email, name}
    share_url  = data.get('share_url', '')
    hod_name   = data.get('hod_name', 'Your HOD')
    dept       = data.get('dept', 'your department')
    semester   = data.get('semester', '')

    if not recipients or not share_url:
        return jsonify({'ok': False, 'error': 'Missing recipients or link'})

    sent = []
    failed = []
    for r in recipients:
        to_email = r.get('email', '').strip()
        to_name  = r.get('name', 'Faculty')
        if not to_email:
            continue

        html = f"""
        <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;background:#f8fafc;padding:32px">
          <div style="background:#0a2540;border-radius:12px;padding:24px;text-align:center;margin-bottom:24px">
            <img src="https://jainuniversity.ac.in/static/images/logo.png" height="44" style="margin-bottom:12px;opacity:.9" onerror="this.style.display='none'">
            <div style="color:#fff;font-size:20px;font-weight:800;letter-spacing:-.02em">Office of Academics</div>
            <div style="color:#94a3b8;font-size:12px;margin-top:4px">Jain (Deemed-to-be University)</div>
          </div>
          <div style="background:#fff;border-radius:12px;padding:28px;border:1px solid #e2e8f0">
            <p style="color:#0a2540;font-size:16px;font-weight:700;margin:0 0 8px">Dear {to_name},</p>
            <p style="color:#475569;font-size:14px;line-height:1.7;margin:0 0 16px">
              <strong>{hod_name}</strong>, Head of <strong>{dept}</strong>, has shared the faculty
              {('Semester Readiness' if 'readiness' in share_url else 'Semester Closure')} Checklist link for
              <strong>{semester}</strong> with you. Please fill it at the earliest.
            </p>
            <div style="background:#f0f9ff;border:1.5px solid #bae6fd;border-radius:8px;padding:16px;margin-bottom:20px">
              <p style="color:#0369a1;font-size:12px;font-weight:700;margin:0 0 8px;text-transform:uppercase;letter-spacing:.05em">Your Faculty Checklist Link</p>
              <a href="{share_url}" style="color:#2563eb;font-size:14px;word-break:break-all">{share_url}</a>
            </div>
            <a href="{share_url}" style="display:inline-block;background:#0a2540;color:#fff;text-decoration:none;padding:12px 28px;border-radius:8px;font-weight:700;font-size:14px">
              Open Checklist &rarr;
            </a>
            <p style="color:#94a3b8;font-size:12px;margin-top:24px">
              This link was shared by your HOD via the Jain University Academic Portal. Please complete your checklist before the deadline.
            </p>
          </div>
        </div>
        """
        ok = _send_email(to_email, f'Faculty Checklist Link — {dept} ({semester})', html)
        if ok:
            sent.append(to_email)
        else:
            failed.append(to_email)

    return jsonify({'ok': True, 'sent': sent, 'failed': failed})



# ═════════════════════════════════════════════════════════════
# HOD SUBMISSION (generates share_token used by faculty link)
# ═════════════════════════════════════════════════════════════

@app.route('/api/submit', methods=['POST'])
def submit():
    data = request.json
    sub_id = data.get('_id')

    # Security check for late submission
    form_type = data.get('form_type', 'readiness')
    settings = get_global_settings()
    deadline = settings.get('readiness_deadline') if form_type == 'readiness' else settings.get('closure_deadline')
    if not session.get('admin') and is_deadline_passed(deadline):
        if not has_access_override(session.get('user_email'), form_type):
            return jsonify({'ok': False, 'error': f'The deadline has passed. Late submissions are blocked. Please request admin access.'})

    if '_id' in data:
        del data['_id']

    data['timestamp'] = datetime.utcnow().isoformat()
    is_draft = data.get('_draft', False)

    if sub_id:
        existing = submissions_col.find_one({'_id': ObjectId(sub_id)})
        if not existing:
            return jsonify({'ok': False, 'error': 'Submission not found'})

        edit_count = existing.get('edit_count', 0)
        edit_request = existing.get('edit_request', {})
        
        # We only count it as an edit if the previous save was NOT a draft
        was_previously_submitted = not existing.get('_draft', False)

        if not is_draft and was_previously_submitted:
            if edit_count >= 2 and edit_request.get('status') != 'approved':
                return jsonify({'ok': False, 'error': 'Edit limit exceeded. Please request admin approval.'})

            if edit_request.get('status') == 'approved':
                data['edit_request'] = None

            data['edit_count'] = edit_count + 1
        else:
            # Saving a draft, or finalizing a draft for the first time keeps current edit_count
            data['edit_count'] = edit_count

        # Preserve share_token if exists
        if existing.get('share_token'):
            data['share_token'] = existing['share_token']
        else:
            data['share_token'] = secrets.token_urlsafe(16)

        submissions_col.update_one({'_id': ObjectId(sub_id)}, {'$set': data})
        return jsonify({'ok': True, 'id': sub_id, 'share_token': data['share_token']})
    else:
        data['edit_count'] = 0
        data['share_token'] = secrets.token_urlsafe(16)
        result = submissions_col.insert_one(data)
        return jsonify({'ok': True, 'id': str(result.inserted_id), 'share_token': data['share_token']})

@app.route('/api/upload-evidence', methods=['POST'])
def upload_evidence():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file part'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'error': 'No selected file'})

    if file:
        filename = secure_filename(file.filename)
        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = f"{ts}_{filename}"

        upload_path = os.path.join('static', 'uploads')
        os.makedirs(upload_path, exist_ok=True)

        file.save(os.path.join(upload_path, filename))
        url = url_for('static', filename=f"uploads/{filename}")
        return jsonify({'ok': True, 'url': url})

    return jsonify({'ok': False, 'error': 'Unknown error'})

@app.route('/api/request-edit/<sub_id>', methods=['POST'])
def request_edit(sub_id):
    data = request.json
    comment = data.get('comment', '').strip()

    submissions_col.update_one(
        {'_id': ObjectId(sub_id)},
        {'$set': {'edit_request': {'pending': True, 'comment': comment, 'status': 'pending', 'timestamp': datetime.utcnow().isoformat()}}}
    )
    return jsonify({'ok': True})

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if not name or not email or not password:
        return jsonify({'ok': False, 'error': 'Name, Email, and Password are required'})

    if len(password) < 7:
        return jsonify({'ok': False, 'error': 'Password must be at least 7 characters'})

    existing = users_col.find_one({'email': email})
    if existing:
        return jsonify({'ok': False, 'error': 'You already have an account with this email. Please use Login.'})

    is_admin = is_admin_email(email)
    users_col.insert_one({
        'name': name,
        'email': email,
        'password': generate_password_hash(password),
        'created_at': datetime.utcnow().isoformat(),
        'first_time_login': True,
        'is_admin': is_admin
    })
    session['user_email'] = email
    session['user_name'] = name

    if is_admin:
        session['admin'] = True
        redirect_url = '/admin'
    else:
        redirect_url = '/'

    return jsonify({'ok': True, 'redirect_url': redirect_url, 'is_admin': is_admin, 'email': email})

@app.route('/dev-login')
def dev_login():
    email = 'santosh.ks@jainuniversity.ac.in'
    session['user_email'] = email
    session['user_name'] = 'santosh'
    if is_admin_email(email):
        session['admin'] = True
    return redirect('/admin' if session.get('admin') else '/')

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({'ok': False, 'error': 'Email and Password are required'})

    user = users_col.find_one({'email': email})
    if not user:
        return jsonify({'ok': False, 'error': 'User not found. Please register as a New User.'})

    if not user.get('password'):
        return jsonify({'ok': False, 'error': 'This account was registered with Google Sign-In. Please sign in with Google or use Forgot Password to set a password.'})

    if not check_password_hash(user['password'], password):
        return jsonify({'ok': False, 'error': 'Invalid credentials'})

    session['user_email'] = user['email']
    session['user_name'] = user['name']

    is_admin = is_admin_email(email)
    if is_admin:
        session['admin'] = True
        redirect_url = '/admin'
    else:
        redirect_url = '/'

    return jsonify({'ok': True, 'redirect_url': redirect_url, 'is_admin': is_admin, 'email': email})

# Removed passcode endpoints

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    email = request.json.get('email', '').strip()
    user = users_col.find_one({'email': email})
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})

    otp = str(random.randint(1000, 9999))
    users_col.update_one({'email': email}, {'$set': {'reset_otp': otp}})

    send_otp_email(email, otp)

    session['pending_email'] = email
    return jsonify({'ok': True})

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    email = session.get('pending_email') or session.get('user_email')
    if not email:
        return jsonify({'ok': False, 'error': 'Session expired. Please try again.'})
    data = request.json
    otp = data.get('otp')
    new_password = data.get('password')

    if not otp or not new_password or len(str(new_password)) < 7:
        return jsonify({'ok': False, 'error': 'Invalid OTP or password must be at least 7 characters'})

    user = users_col.find_one({'email': email})
    if str(user.get('reset_otp')) != str(otp):
        return jsonify({'ok': False, 'error': 'Invalid OTP'})

    users_col.update_one({'email': email}, {'$set': {'password': generate_password_hash(new_password), 'reset_otp': None}})
    session['user_email'] = email
    session['user_name'] = user['name']
    session.pop('pending_email', None)
    return jsonify({'ok': True})

@app.route('/api/update-profile', methods=['POST'])
def update_profile():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    dept = request.json.get('department')
    users_col.update_one({'email': session['user_email']}, {'$set': {
        'department': dept,
        'first_time_login': False
    }})
    return jsonify({'ok': True})

@app.route('/api/complete-tour', methods=['POST'])
def complete_tour():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    users_col.update_one({'email': session['user_email']}, {'$set': {
        'first_time_login': False
    }})
    return jsonify({'ok': True})

@app.route('/api/logout', methods=['POST', 'GET'])
def logout():
    session.pop('user_email', None)
    session.pop('user_name', None)
    if request.method == 'GET':
        return redirect(url_for('login'))
    return jsonify({'ok': True})

@app.route('/api/my-submissions')
def my_submissions():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})

    email = session['user_email'].strip()
    user_doc = users_col.find_one({'email': email}) or {}
    dept = user_doc.get('department', '')
    form_type = request.args.get('type')
    show_all = (request.args.get('all') == 'true')

    if show_all and (session.get('admin') or is_admin_email(email)):
        query = {}
    elif dept:
        query = {'$or': [{'identity.submitterEmail': email}, {'identity.dept': dept}]}
    else:
        query = {'identity.submitterEmail': email}

    if form_type:
        query['form_type'] = form_type

    docs = list(submissions_col.find(query, sort=[('timestamp', -1)]))
    for d in docs:
        d['_id'] = str(d['_id'])
        # Attach faculty submission count
        d['faculty_submission_count'] = faculty_submissions_col.count_documents({'parent_submission_id': d['_id']})
    return jsonify({'ok': True, 'submissions': docs})

@app.route('/api/get-submission/<sub_id>')
def get_submission(sub_id):
    """Get a single HOD submission with its faculty submissions (for share/review page)."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    try:
        doc = submissions_col.find_one({'_id': ObjectId(sub_id)})
        if not doc:
            return jsonify({'ok': False, 'error': 'Not found'})
        doc['_id'] = str(doc['_id'])
        fac_docs = list(faculty_submissions_col.find({'parent_submission_id': sub_id}, sort=[('timestamp', -1)]))
        for f in fac_docs:
            f['_id'] = str(f['_id'])
        doc['_faculty_submissions'] = fac_docs
        return jsonify({'ok': True, 'submission': doc})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/export/<submission_id>')
def export_submission(submission_id):
    """Export the full submission (HOD + all faculty) as Excel — admin/HOD download."""
    try:
        doc = submissions_col.find_one({'_id': ObjectId(submission_id)})
    except:
        return "Not found", 404
    if not doc:
        return "Not found", 404
    fac_docs = list(faculty_submissions_col.find({'parent_submission_id': submission_id}))
    doc['_faculty_submissions'] = fac_docs
    wb = build_workbook([doc])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dept = doc.get('identity', {}).get('dept', 'Department').replace(' ', '_')[:30]
    form_type = doc.get('form_type', 'readiness').title()
    return send_file(buf, as_attachment=True,
                     download_name=f"Sem{form_type}_{dept}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/api/export-hod-checklist/<submission_id>')
def export_hod_checklist(submission_id):
    """Download the HOD's filled checklist as an Excel template
    that the HOD can hand to faculty as a reference."""
    try:
        doc = submissions_col.find_one({'_id': ObjectId(submission_id)})
    except:
        return "Not found", 404
    if not doc:
        return "Not found", 404

    form_type = doc.get('form_type', 'readiness')
    wb = build_hod_checklist_only(doc, form_type)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dept = doc.get('identity', {}).get('dept', 'Department').replace(' ', '_')[:30]
    label = 'Closure' if form_type == 'closure' else 'Readiness'
    return send_file(buf, as_attachment=True,
                     download_name=f"HOD_{label}_{dept}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Admin ────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('admin'):
        return redirect(url_for('admin_dashboard'))

    if request.method == 'GET' and request.args.get('reset') == '1':
        session.pop('admin_otp', None)
        session.pop('admin_otp_sent', None)
        session.pop('admin_otp_time', None)
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        if session.get('admin_otp_sent'):
            entered_otp = request.form.get('otp', '').strip()
            session_otp = session.get('admin_otp')
            otp_time = session.get('admin_otp_time', 0)
            
            is_expired = (datetime.utcnow().timestamp() - otp_time) > 600
            if is_expired:
                flash('OTP has expired. Please try again.')
                session.pop('admin_otp', None)
                session.pop('admin_otp_sent', None)
                session.pop('admin_otp_time', None)
                return render_template('admin_login.html')

            if entered_otp and session_otp and entered_otp == session_otp:
                session['admin'] = True
                session.pop('admin_otp', None)
                session.pop('admin_otp_sent', None)
                session.pop('admin_otp_time', None)
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid OTP')
                return render_template('admin_login.html')
        else:
            u = request.form.get('username')
            p = request.form.get('password')
            if u == ADMIN_USERNAME and (p == 'admin' or p == ADMIN_PASSWORD):
                otp = str(random.randint(100000, 999999))
                session['admin_otp'] = otp
                session['admin_otp_time'] = datetime.utcnow().timestamp()
                session['admin_otp_sent'] = True
                
                email_sent = send_admin_otp_email('santosh.ks@jainuniversity.ac.in', otp)
                if not email_sent:
                    flash('Failed to send OTP email. Please check SMTP settings.')
                    session.pop('admin_otp', None)
                    session.pop('admin_otp_sent', None)
                    session.pop('admin_otp_time', None)
                    return render_template('admin_login.html')
                
                flash('A 6-digit OTP has been sent to santosh.ks@jainuniversity.ac.in')
                return render_template('admin_login.html')
            else:
                flash('Invalid credentials')
                return render_template('admin_login.html')

    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/impersonate/<email>')
def admin_impersonate(email):
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    user = users_col.find_one({'email': email})
    if not user:
        return "User not found", 404
    session['user_email'] = user['email']
    session['user_name'] = user['name']
    return redirect('/')

@app.route('/admin')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))

    # Shared defaults, so the IEA panel is populated exactly like readiness/closure
    settings = get_global_settings()
    dynamic_admins = set(e.strip().lower() for e in settings.get('admin_emails', []) if e)

    submissions = list(submissions_col.find().sort('timestamp', -1))

    # Count faculty responses for every submission in one grouped query instead of
    # one count_documents per row — that N+1 was what pushed this page past the
    # gunicorn worker timeout on large datasets.
    faculty_counts = {}
    for row in faculty_submissions_col.aggregate([
        {'$group': {'_id': '$parent_submission_id', 'count': {'$sum': 1}}}
    ]):
        faculty_counts[row['_id']] = row['count']

    for s in submissions:
        s['_id'] = str(s['_id'])
        s['faculty_submission_count'] = faculty_counts.get(s['_id'], 0)
        # The dashboard (and its JS) reads s.identity.* everywhere — one legacy row
        # without an identity block would otherwise take down the whole page.
        if not isinstance(s.get('identity'), dict):
            s['identity'] = {}

    # Latest department per submitter, derived from the submissions already loaded
    # (they are sorted newest-first, so the first hit for an email is the latest).
    latest_dept_by_email = {}
    for s in submissions:
        identity = s.get('identity') or {}
        email = (identity.get('submitterEmail') or '').strip()
        if email and email not in latest_dept_by_email and identity.get('dept'):
            latest_dept_by_email[email] = identity['dept']

    today_str = datetime.utcnow().strftime('%Y-%m-%d')
    logged_in_today_count = 0
    users = list(users_col.find().sort('created_at', -1))
    for u in users:
        u['_id'] = str(u['_id'])
        email = (u.get('email') or '').strip()
        # Resolved from data already in hand — is_admin_email() would fire two extra
        # queries per user, which is the other half of the old page-load cost.
        u['is_admin'] = bool(email) and (
            email.lower() in ADMIN_EMAILS
            or email.lower() in dynamic_admins
            or u.get('is_admin') is True
            or u.get('role') == 'admin'
        )
        u['can_readiness'] = u.get('can_readiness', True)
        u['can_closure'] = u.get('can_closure', True)
        u['can_iea'] = u.get('can_iea', True)
        last_log = u.get('last_login', '')
        is_today = bool(last_log and last_log.startswith(today_str))
        u['logged_in_today'] = is_today
        if is_today:
            logged_in_today_count += 1
        u['latest_dept'] = u.get('department') or latest_dept_by_email.get(email) or 'N/A'

    iea_subs = [_iea_public_doc(d) for d in
                iea_col.find().sort([('school', 1), ('department', 1), ('level', 1)])]

    access_requests = list(db['access_requests'].find().sort('timestamp', -1))
    for r in access_requests:
        r['_id'] = str(r['_id'])
        mod = module_from_label(r.get('module'))
        r['module_key'] = mod['key'] if mod else ''


    pwa_stats = db['pwa_analytics'].find_one({'_id': 'pwa_stats'}) or {'installs': 0, 'launches': 0}
    
    notifications = list(db['notifications'].find().sort('created_at', -1).limit(10))
    for n in notifications:
        n['_id'] = str(n['_id'])
    
    return render_template('admin.html',
                           submissions=submissions,
                           hod_sections=HOD_SECTIONS,
                           faculty_sections=FACULTY_SECTIONS,
                           users=users, settings=settings, departments=DEPARTMENTS,
                           access_requests=access_requests,
                           pwa_stats=pwa_stats,
                           notifications=notifications,
                           logged_in_today_count=logged_in_today_count,
                           modules=[MODULES[k] for k in MODULE_ORDER],
                           iea_submissions=iea_subs,
                           iea_years=IEA_YEARS,
                           iea_sections=IEA_SECTIONS)

@app.route('/admin/update-user-permissions', methods=['POST'])
def update_user_permissions():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    if not email:
        return jsonify({'ok': False, 'error': 'Email is required'})

    can_readiness = bool(data.get('can_readiness', True))
    can_closure = bool(data.get('can_closure', True))
    can_iea = bool(data.get('can_iea', True))

    users_col.update_one(
        {'email': email},
        {'$set': {
            'can_readiness': can_readiness,
            'can_closure': can_closure,
            'can_iea': can_iea
        }},
        upsert=True
    )
    return jsonify({'ok': True, 'email': email, 'can_readiness': can_readiness, 'can_closure': can_closure, 'can_iea': can_iea})

@app.route('/admin/settings', methods=['POST'])
def save_settings():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json or {}

    prev_settings = get_global_settings()

    update = {
        'enabled_years': data.get('enabled_years', ['2024-25', '2025-26', '2026-27', '2027-28']),
        'enabled_semesters': data.get('enabled_semesters', ['Even', 'Odd'])
    }
    # Keep the previous value when the panel does not send a module's field, so a
    # partial save can never silently switch a module off.
    for module_key in MODULE_ORDER:
        mod = MODULES[module_key]
        update[mod['enabled_key']] = bool(data.get(mod['enabled_key'],
                                                   prev_settings.get(mod['enabled_key'], True)))
        update[mod['deadline_key']] = data.get(mod['deadline_key'],
                                               prev_settings.get(mod['deadline_key'], ''))

    settings_col.update_one({'_id': 'global'}, {'$set': update}, upsert=True)

    # Broadcast a notification when a module is opened or its deadline moves —
    # identical treatment for readiness, closure and IEA.
    notif_title = ""
    notif_body = ""
    for module_key in MODULE_ORDER:
        mod = MODULES[module_key]
        was_enabled = bool(prev_settings.get(mod['enabled_key'], False))
        is_enabled = update[mod['enabled_key']]
        prev_dl = prev_settings.get(mod['deadline_key'], '')
        new_dl = update[mod['deadline_key']]

        if not was_enabled and is_enabled:
            notif_title = f"{mod['label']} Added"
            notif_body = f"{mod['label']} has been added! Please check your portal."
            break
        if new_dl and new_dl != prev_dl:
            notif_title = f"{mod['label']} Deadline Extended"
            notif_body = f"{mod['label']} last date has been updated/extended to {new_dl}!"
            break

    if notif_title:
        db['notifications'].insert_one({
            'title': notif_title,
            'body': notif_body,
            'created_at': datetime.utcnow().isoformat(),
            'type': 'broadcast'
        })

    return jsonify({'ok': True})

@app.route('/admin/toggle-admin-access', methods=['POST'])
def toggle_admin_access():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json or {}
    target_email = data.get('email', '').strip().lower()
    make_admin = data.get('make_admin', True)

    if not target_email:
        return jsonify({'ok': False, 'error': 'Email is required'})

    # Update in users collection
    users_col.update_one({'email': target_email}, {'$set': {'is_admin': make_admin}}, upsert=True)

    # Update in global settings admin_emails list
    global_settings = settings_col.find_one({'_id': 'global'}) or {}
    admin_list = set(e.lower() for e in global_settings.get('admin_emails', []))
    if make_admin:
        admin_list.add(target_email)
    else:
        admin_list.discard(target_email)

    settings_col.update_one(
        {'_id': 'global'},
        {'$set': {'admin_emails': list(admin_list)}},
        upsert=True
    )

    return jsonify({'ok': True, 'email': target_email, 'is_admin': make_admin})

@app.route('/admin/create-user', methods=['POST'])
def admin_create_user():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    department = data.get('department', '').strip()
    role = data.get('role', 'user').strip().lower()
    password = data.get('password', '')

    if not email or not name:
        return jsonify({'ok': False, 'error': 'Name and Email ID are required'})

    is_admin = (role == 'admin')
    update_data = {
        'name': name,
        'email': email,
        'department': department,
        'is_admin': is_admin,
        'role': role,
        'first_time_login': False if department else True
    }

    if password:
        update_data['password'] = generate_password_hash(password)

    existing = users_col.find_one({'email': email})
    if not existing:
        update_data['created_at'] = datetime.utcnow().isoformat()
        update_data['auth_provider'] = 'admin_created'
        users_col.insert_one(update_data)
    else:
        users_col.update_one({'email': email}, {'$set': update_data})

    if is_admin:
        global_settings = settings_col.find_one({'_id': 'global'}) or {}
        admin_list = set(e.lower() for e in global_settings.get('admin_emails', []))
        admin_list.add(email)
        settings_col.update_one({'_id': 'global'}, {'$set': {'admin_emails': list(admin_list)}}, upsert=True)

    return jsonify({'ok': True})

@app.route('/admin/delete-user/<user_id>', methods=['POST'])
def admin_delete_user(user_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    try:
        user = users_col.find_one({'_id': ObjectId(user_id)})
        if not user:
            user = users_col.find_one({'email': user_id})
        if user:
            users_col.delete_one({'_id': user['_id']})
            # Remove from admin_emails list if present
            global_settings = settings_col.find_one({'_id': 'global'}) or {}
            admin_list = set(e.lower() for e in global_settings.get('admin_emails', []))
            admin_list.discard(user.get('email', '').lower())
            settings_col.update_one({'_id': 'global'}, {'$set': {'admin_emails': list(admin_list)}}, upsert=True)
            return jsonify({'ok': True})
        return jsonify({'ok': False, 'error': 'User not found'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/get-notifications', methods=['GET'])
def get_notifications():
    notifs = list(db['notifications'].find().sort('created_at', -1).limit(10))
    for n in notifs:
        n['_id'] = str(n['_id'])
    return jsonify({'ok': True, 'notifications': notifs})

@app.route('/api/send-notification', methods=['POST'])
def send_custom_notification():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json or {}
    title = data.get('title', '').strip()
    body = data.get('body', '').strip()
    if not title or not body:
        return jsonify({'ok': False, 'error': 'Missing title or body'})
    
    db['notifications'].insert_one({
        'title': title,
        'body': body,
        'created_at': datetime.utcnow().isoformat(),
        'type': 'custom'
    })
    return jsonify({'ok': True})

@app.route('/admin/delete-notification/<notif_id>', methods=['POST'])
def delete_notification(notif_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    from bson import ObjectId
    db['notifications'].delete_one({'_id': ObjectId(notif_id)})
    return jsonify({'ok': True})

@app.route('/api/request-extension', methods=['POST'])
def request_extension():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    data = request.json or {}
    comment = data.get('comment', '').strip()
    mod = module_from_label(data.get('module', ''))
    if not mod:
        return jsonify({'ok': False, 'error': 'Unknown module for this request'})
    module = mod['label']  # store the canonical label so admin approval can map it back

    db['access_requests'].update_one(
        {
            'user_email': session['user_email'],
            'module': module
        },
        {
            '$set': {
                'user_name': session['user_name'],
                'comment': comment,
                'status': 'pending',
                'timestamp': datetime.utcnow().isoformat()
            }
        },
        upsert=True
    )
    return jsonify({'ok': True})

def send_override_approval_email(to_email, name, module_name, admin_comment):
    login_url = get_base_url() + "login"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #ecfdf5; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); border-top: 6px solid #10b981;">
          <h2 style="color: #065f46; margin-bottom: 20px;">✅ Late Submission Access Approved</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear HOD {name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">The Office of Academics administration has reviewed and <strong>approved</strong> your request for late submission access to <strong>{module_name}</strong>.</p>

          <div style="margin-bottom: 25px; background: #f0fdf4; border-left: 4px solid #10b981; padding: 12px 15px; border-radius: 6px; text-align: left;">
            <strong style="color: #065f46; font-size: 13px; display: block; margin-bottom: 6px;">Administrator Remarks:</strong>
            <span style="color: #374151; font-size: 14px; font-style: italic;">"{admin_comment or 'No comment provided.'}"</span>
          </div>

          <p style="color: #3d4460; font-size: 14px; margin-bottom: 25px;">You may now log in and complete your submission form.</p>

          <div style="margin-bottom: 30px;">
            <a href="{login_url}" style="background-color: #10b981; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Go to Submission Portal</a>
          </div>

          <p style="color: #8892aa; font-size: 12px;">This is an automated notification. Please do not reply directly to this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, f'✅ APPROVED: Late Submission Access Granted for {module_name}', html_content)

@app.route('/admin/approve-access/<request_id>', methods=['POST'])
def approve_access(request_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    req = db['access_requests'].find_one({'_id': ObjectId(request_id)})
    if not req:
        return jsonify({'ok': False, 'error': 'Request not found'})
    
    data = request.json or {}
    admin_comment = data.get('admin_comment', '').strip()
    
    email = req['user_email']
    module = req.get('module', '')
    mod = module_from_label(module)
    if not mod:
        return jsonify({'ok': False, 'error': f'Unrecognised module "{module}" on this request'})

    users_col.update_one({'email': email}, {'$set': {mod['override_field']: True}})
    db['access_requests'].update_one(
        {'_id': ObjectId(request_id)}, 
        {'$set': {'status': 'approved', 'admin_comment': admin_comment}}
    )
    
    # Send email notification to HOD
    send_override_approval_email(email, req.get('user_name', 'HOD'), module, admin_comment)
    
    return jsonify({'ok': True})

@app.route('/admin/revoke-access/<email>/<module_type>', methods=['POST'])
def revoke_access(email, module_type):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    mod = get_module(module_type) or module_from_label(module_type)
    if not mod:
        return jsonify({'ok': False, 'error': f'Unknown module "{module_type}"'})
    users_col.update_one({'email': email}, {'$set': {mod['override_field']: False}})
    db['access_requests'].delete_many({'user_email': email, 'module': mod['label']})
    return jsonify({'ok': True})

@app.route('/admin/send-manual-reminders', methods=['POST'])
def send_manual_reminders():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})

    settings = get_global_settings()
    sent_count = 0

    # One pass per module — readiness, closure and IEA are all handled the same
    for module_key in MODULE_ORDER:
        mod = MODULES[module_key]
        deadline, date_str = get_module_deadline(settings, module_key)
        if not deadline:
            continue
        for user in users_col.find():
            email = user.get('email')
            if not email:
                continue
            name = user.get('name') or 'HOD'
            if has_module_submission(module_key, email):
                continue
            print(f"Manual reminder: Sending {module_key} to {name} ({email})...")
            if send_deadline_reminder_email(email, name, mod['label'], date_str):
                sent_count += 1

    return jsonify({'ok': True, 'sent_count': sent_count})

@app.route('/admin/send-targeted-reminders', methods=['POST'])
def send_targeted_reminders():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})

    data = request.json or {}
    target_type = data.get('target_type')
    department = data.get('department')

    if not target_type:
        return jsonify({'ok': False, 'error': 'Missing target_type'})

    settings = get_global_settings()
    sent_count = 0

    date_strs = {k: get_module_deadline(settings, k)[1] for k in MODULE_ORDER}

    def remind(user, module_key):
        """Email this HOD about the module if their submission is still pending."""
        nonlocal sent_count
        email = user.get('email')
        if not email:
            return
        name = user.get('name') or 'HOD'
        if has_module_submission(module_key, email):
            return
        mod = MODULES[module_key]
        print(f"Targeted reminder: Sending {module_key} to {name} ({email})...")
        if send_deadline_reminder_email(email, name, mod['label'], date_strs[module_key]):
            sent_count += 1

    # 'pending_readiness' / 'pending_closure' / 'pending_iea' target one module
    single_module = None
    if target_type.startswith('pending_'):
        single_module = get_module(target_type[len('pending_'):])
        if not single_module:
            return jsonify({'ok': False, 'error': f'Unknown target type "{target_type}"'})

    if single_module:
        for user in users_col.find():
            remind(user, single_module['key'])

    elif target_type == 'all_pending':
        for user in users_col.find():
            for module_key in MODULE_ORDER:
                remind(user, module_key)

    elif target_type == 'specific_dept':
        if not department:
            return jsonify({'ok': False, 'error': 'Missing department for specific_dept target group'})

        found_any = False
        for user in users_col.find():
            user_dept = user.get('department')
            if not user_dept:
                latest_sub = submissions_col.find_one({'identity.submitterEmail': user.get('email')},
                                                      sort=[('timestamp', -1)])
                if latest_sub:
                    user_dept = (latest_sub.get('identity') or {}).get('dept')

            if user_dept == department:
                found_any = True
                for module_key in MODULE_ORDER:
                    remind(user, module_key)

        if not found_any:
            return jsonify({'ok': False, 'error': f'No registered HOD account found for department "{department}"'})

    else:
        return jsonify({'ok': False, 'error': f'Unknown target type "{target_type}"'})

    return jsonify({'ok': True, 'sent_count': sent_count})

@app.route('/admin/approve-edit/<sub_id>', methods=['POST'])
def approve_edit(sub_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})

    submissions_col.update_one(
        {'_id': ObjectId(sub_id)},
        {'$set': {'edit_request.status': 'approved', 'edit_request.pending': False}}
    )
    return jsonify({'ok': True})

@app.route('/admin/export-all')
def export_all():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    docs = list(submissions_col.find().sort('timestamp', -1))
    for d in docs:
        fac_docs = list(faculty_submissions_col.find({'parent_submission_id': str(d['_id'])}))
        d['_faculty_submissions'] = fac_docs
    wb = build_workbook(docs)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"SemReadiness_ALL_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/delete/<sid>', methods=['POST'])
def delete_submission(sid):
    if not session.get('admin'):
        return jsonify({'ok': False})
    submissions_col.delete_one({'_id': ObjectId(sid)})
    # Also delete linked faculty submissions
    faculty_submissions_col.delete_many({'parent_submission_id': sid})
    return jsonify({'ok': True})

# Allow the HOD (owner) to also delete their own submission
@app.route('/api/delete-submission/<sid>', methods=['POST'])
def delete_my_submission(sid):
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    try:
        doc = submissions_col.find_one({'_id': ObjectId(sid)})
        if not doc:
            return jsonify({'ok': False, 'error': 'Not found'})
        # Owner-only unless admin
        if not session.get('admin') and doc.get('identity', {}).get('submitterEmail') != session.get('user_email'):
            return jsonify({'ok': False, 'error': 'Not authorised'})
        submissions_col.delete_one({'_id': ObjectId(sid)})
        faculty_submissions_col.delete_many({'parent_submission_id': sid})
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ═════════════════════════════════════════════════════════════
# WORKBOOK BUILDERS
# ═════════════════════════════════════════════════════════════

def _styles():
    hdr_fill = PatternFill("solid", fgColor="0A2558")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    gold_fill = PatternFill("solid", fgColor="F4A819")
    gold_font = Font(color="1A1A1A", bold=True, size=10, name="Calibri")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return hdr_fill, hdr_font, gold_fill, gold_font, thin, center


def build_hod_checklist_only(doc, form_type='readiness'):
    """Build a single-sheet Excel that the HOD can download and share with faculty.
    Contains: identity header + HOD's filled checklist + a blank Faculty column for
    each faculty (or simply the checklist for reference).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill, hdr_font, gold_fill, gold_font, thin, center = _styles()

    label = 'Closure' if form_type == 'closure' else 'Readiness'
    ws = wb.create_sheet(f"HOD {label} Checklist")
    idt = doc.get('identity', {})

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1, value=f"JAIN University — HOD Semester {label} Submission")
    t.font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    t.fill = hdr_fill
    t.alignment = center
    ws.row_dimensions[1].height = 32

    # Identity block
    identity_rows = [
        ("Department",       idt.get('dept', '')),
        ("Campus",           idt.get('campus', '')),
        ("HOD Name",         idt.get('hodName', '') or idt.get('hod', '')),
        ("HOD Email",        idt.get('hodEmail', '') or idt.get('email', '')),
        ("Semester",         idt.get('semester', '') or idt.get('sem', '')),
        ("Academic Year",    idt.get('acYear', '')),
        ("Submission Date",  idt.get('subDate', '') or idt.get('date', '')),
        ("Faculty Deadline", idt.get('hodDeadline', '')),
        ("Mid Sem Review Date", doc.get('hodChecklist', {}).get('28', {}).get('status', '')),
    ]
    r = 3
    for label_, val in identity_rows:
        c1 = ws.cell(row=r, column=1, value=label_)
        c1.fill = gold_fill
        c1.font = gold_font
        c1.border = thin
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = Font(name="Calibri", size=10)
        c2.border = thin
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1

    r += 1

    # Checklist / Report headers
    if form_type == 'closure':
        headers = ['#', 'Vertical (HOD Report Section)', 'HOD Narrative', '', '']
    else:
        headers = ['#', 'Section', 'Checklist Item', 'Status', 'Remark / Notes']

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[r].height = 28
    r += 1

    if form_type == 'closure':
        hod_rpt = doc.get('hodReport', {})
        for item in HOD_CLOSURE_SECTIONS:
            val = hod_rpt.get(item['id'], {})
            if isinstance(val, dict):
                row_vals = [item['id'], item['title'], val.get('text', ''), val.get('link', ''), val.get('file_url', '')]
            else:
                row_vals = [item['id'], item['title'], str(val), '', '']
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
    else:
        chk = doc.get('hodChecklist', {})
        for sec in HOD_SECTIONS:
            sec_title = sec['title']
            for item in sec['items']:
                val = chk.get(item['id'], {})
                row_vals = [item['id'], sec_title, item['text'], val.get('status', ''), val.get('remark', '')]
                for c, v in enumerate(row_vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if c == 4:
                        if v == 'No':
                            cell.fill = PatternFill("solid", fgColor="FFE8E8")
                        elif v == 'Yes':
                            cell.fill = PatternFill("solid", fgColor="E8F8EE")
                r += 1

    for i, w in enumerate([8, 30, 55, 14, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Second sheet: a blank Faculty Checklist template (for faculty reference) ──
    _, fac_sections = _sections_for(form_type)
    ws2 = wb.create_sheet("Faculty Checklist (Template)")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t2 = ws2.cell(row=1, column=1, value=f"Faculty Semester {label} Checklist — Auto-filled from HOD")
    t2.font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
    t2.fill = hdr_fill
    t2.alignment = center
    ws2.row_dimensions[1].height = 28

    fac_id_rows = [
        ("Department",     idt.get('dept', '')),
        ("Campus",         idt.get('campus', '')),
        ("HOD Name",       idt.get('hodName', '') or idt.get('hod', '')),
        ("Semester",       idt.get('semester', '') or idt.get('sem', '')),
        ("Academic Year",  idt.get('acYear', '')),
        ("Faculty Name",   ""),
        ("Faculty Email",  ""),
        ("Course Name",    ""),
        ("Course Code",    ""),
    ]
    r = 3
    for label_, val in fac_id_rows:
        c1 = ws2.cell(row=r, column=1, value=label_)
        c1.fill = gold_fill
        c1.font = gold_font
        c1.border = thin
        c2 = ws2.cell(row=r, column=2, value=val)
        c2.font = Font(name="Calibri", size=10)
        c2.border = thin
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1

    headers2 = ['#', 'Faculty Checklist Item', 'Status (Yes/No/N/A)', 'Remarks / Evidence Link']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
    ws2.row_dimensions[r].height = 28
    r += 1

    for sec in fac_sections:
        # section header row
        cell = ws2.cell(row=r, column=1, value=sec['title'])
        cell.fill = gold_fill
        cell.font = gold_font
        cell.border = thin
        cell.alignment = Alignment(horizontal='left')
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        for item in sec['items']:
            row_vals = [item['id'], item['text'], '', '']
            for c, v in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
    for i, w in enumerate([8, 60, 18, 36], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    return wb


def build_workbook(submissions):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill, hdr_font, gold_fill, gold_font, thin, center = _styles()

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[row].height = 36

    def add_title(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        t = ws.cell(row=1, column=1, value=text)
        t.font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
        t.fill = hdr_fill
        t.alignment = center
        ws.row_dimensions[1].height = 28

    # Sheet 1 — Dashboard
    ws1 = wb.create_sheet("Dashboard")
    add_title(ws1, "JAIN University — Semester Submissions Dashboard", 20)
    heads = ['Timestamp', 'Form Type', 'Campus', 'Department', 'HOD Name', 'HOD Email', 'Semester', 'Acad Year', 'Sub Date', 'HOD Deadline', 'Mid Sem Date',
             'Faculty Submissions', 'HOD Checklist Yes/Total']
    style_header(ws1, heads, 2)
    for r, d in enumerate(submissions, 3):
        idt = d.get('identity', {})
        chk = d.get('hodChecklist', {})
        yes_count = sum(1 for v in chk.values() if isinstance(v, dict) and v.get('status') == 'Yes')
        fac_subs = len(d.get('_faculty_submissions', []))
        vals = [d.get('timestamp', ''), d.get('form_type', 'readiness'), idt.get('campus', ''),
                idt.get('dept', ''), idt.get('hodName', '') or idt.get('hod', ''),
                idt.get('hodEmail', '') or idt.get('email', ''),
                idt.get('semester', '') or idt.get('sem', ''), idt.get('acYear', ''),
                idt.get('subDate', '') or idt.get('date', ''), idt.get('hodDeadline', ''), chk.get('28', {}).get('status', ''),
                fac_subs, f"{yes_count}/{len(chk) if chk else 0}"]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor="F5F8FF" if r % 2 == 0 else "FFFFFF")
    ws1.column_dimensions['A'].width = 20
    for i in range(2, 21):
        ws1.column_dimensions[get_column_letter(i)].width = 16

    # Sheet 2 — Programs
    ws2 = wb.create_sheet("Program Breakdown")
    add_title(ws2, "Program-wise Breakdown", 7)
    style_header(ws2, ['Department', 'Campus', 'Program Name', 'Courses', 'Students', 'Faculty', 'Coordinator'], 2)
    r = 3
    for d in submissions:
        for p in d.get('programs', []):
            vals = [d.get('identity', {}).get('dept', ''), d.get('identity', {}).get('campus', ''),
                    p.get('name', ''), p.get('courses', ''), p.get('students', ''),
                    p.get('faculty', ''), p.get('coord', '')]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            r += 1
    for i, w in enumerate([28, 14, 32, 10, 10, 10, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3 — Faculty Submissions
    ws3 = wb.create_sheet("Faculty Submissions")
    add_title(ws3, "Faculty Submissions (via Shared Link)", 13)
    style_header(ws3, ['Form Type', 'Department', 'Campus', 'Semester', 'Faculty Name', 'Faculty Email',
                       'Course', 'Code', 'Program', 'Year/Sem', 'Students', 'HOD Remarks', 'Review Status'], 2)
    r = 3
    for d in submissions:
        for f in d.get('_faculty_submissions', []):
            vals = [f.get('form_type', d.get('form_type', 'readiness')),
                    f.get('dept', ''), f.get('campus', ''), f.get('semester', ''),
                    f.get('faculty_name', ''), f.get('faculty_email', ''),
                    f.get('course_name', ''), f.get('course_code', ''),
                    f.get('program', ''), f.get('year_sem', ''), f.get('no_of_students', ''),
                    f.get('hod_remarks', ''), f.get('hod_review_status', '')]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            r += 1
    for i, w in enumerate([12, 26, 12, 16, 22, 26, 24, 12, 18, 12, 10, 28, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4 — Faculty Checklist Detail
    ws4 = wb.create_sheet("Faculty Checklist Detail")
    add_title(ws4, "Faculty Checklist — Item-wise", 9)
    style_header(ws4, ['Form Type', 'Department', 'Faculty Name', 'Course', 'Code', 'Item #', 'Checklist Item', 'Status', 'Remark'], 2)
    r = 3
    for d in submissions:
        for f in d.get('_faculty_submissions', []):
            chk = f.get('checklist', {})
            ftype = f.get('form_type', d.get('form_type', 'readiness'))
            _, fac_sections = _sections_for(ftype)
            for sec in fac_sections:
                for item in sec['items']:
                    val = chk.get(item['id'], {})
                    if not isinstance(val, dict):
                        val = {}
                    vals = [ftype, f.get('dept', ''), f.get('faculty_name', ''), f.get('course_name', ''),
                            f.get('course_code', ''), item['id'], item['text'],
                            val.get('status', ''), val.get('remark', '')]
                    for c, v in enumerate(vals, 1):
                        cell = ws4.cell(row=r, column=c, value=v)
                        cell.border = thin
                        cell.font = Font(name="Calibri", size=10)
                        if val.get('status') == 'No':
                            cell.fill = PatternFill("solid", fgColor="FFE8E8")
                        elif val.get('status') == 'Yes':
                            cell.fill = PatternFill("solid", fgColor="E8F8EE")
                    r += 1
    for i, w in enumerate([12, 26, 22, 24, 12, 8, 48, 10, 30], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Sheet 5 — HOD Checklist (Readiness)
    ws5 = wb.create_sheet("HOD Checklist")
    add_title(ws5, "HOD Checklist — All Departments", 8)
    style_header(ws5, ['Department', 'Campus', 'Semester', 'Item #', 'Checklist Item', 'Section', 'Status', 'Remark'], 2)
    r = 3
    for d in submissions:
        if d.get('form_type', 'readiness') != 'readiness':
            continue
        chk = d.get('hodChecklist', {})
        for sec in HOD_SECTIONS:
            for item in sec['items']:
                val = chk.get(item['id'], {})
                if not isinstance(val, dict):
                    val = {}
                vals = [d.get('identity', {}).get('dept', ''), d.get('identity', {}).get('campus', ''),
                        d.get('identity', {}).get('semester', ''), item['id'], item['text'],
                        sec['title'].replace('Section ', '').split(': ', 1)[-1],
                        val.get('status', ''), val.get('remark', '')]
                for c, v in enumerate(vals, 1):
                    cell = ws5.cell(row=r, column=c, value=v)
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=10)
                    if val.get('status') == 'No':
                        cell.fill = PatternFill("solid", fgColor="FFE8E8")
                    elif val.get('status') == 'Yes':
                        cell.fill = PatternFill("solid", fgColor="E8F8EE")
                r += 1
    for i, w in enumerate([26, 12, 14, 8, 52, 24, 10, 30], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # Sheet 6 — HOD Closure Report
    ws_c = wb.create_sheet("HOD Closure Reports")
    add_title(ws_c, "HOD Semester Closure — Narrative Reports", 5)
    style_header(ws_c, ['Department', 'Campus', 'Semester', 'Vertical', 'HOD Narrative'], 2)
    r = 3
    for d in submissions:
        if d.get('form_type') != 'closure':
            continue
        rpt = d.get('hodReport', {})
        idt = d.get('identity', {})
        for item in HOD_CLOSURE_SECTIONS:
            val = rpt.get(item['id'], {})
            if isinstance(val, dict):
                text = val.get('text', '')
                link = val.get('link', '')
                file_url = val.get('file_url', '')
                if link or file_url:
                    parts = [text]
                    if link:
                        parts.append(f"[Link: {link}]")
                    if file_url:
                        parts.append(f"[File: {file_url}]")
                    text = " \n".join(parts)
            else:
                text = str(val)
            vals = [idt.get('dept', ''), idt.get('campus', ''),
                    idt.get('semester', '') or idt.get('sem', ''),
                    f"{item['id']}. {item['title']}", text]
            for c, v in enumerate(vals, 1):
                cell = ws_c.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
    for i, w in enumerate([26, 12, 14, 38, 80], 1):
        ws_c.column_dimensions[get_column_letter(i)].width = w

    # Sheet 7 — Gap Analysis (Readiness only)
    ws6 = wb.create_sheet("Gap Analysis")
    add_title(ws6, "Gap Analysis — Readiness Checklist Completion by Department", 9)
    style_header(ws6, ['Department', 'Campus', 'Semester', 'Yes', 'No', 'N/A', 'Blank', 'Completion %', 'Gap Items (No)'], 2)
    all_items = [item for sec in HOD_SECTIONS for item in sec['items']]
    r = 3
    for d in submissions:
        if d.get('form_type', 'readiness') != 'readiness':
            continue
        chk = d.get('hodChecklist', {})
        vals_list = [v for v in chk.values() if isinstance(v, dict)]
        yes = sum(1 for x in vals_list if x.get('status') == 'Yes')
        no = sum(1 for x in vals_list if x.get('status') == 'No')
        na = sum(1 for x in vals_list if x.get('status') == 'N/A')
        blank = sum(1 for x in vals_list if not x.get('status'))
        denom = (yes + no + blank)
        pct = f"{round(yes / denom * 100)}%" if denom else "—"
        gaps = "; ".join(f"{i['id']}. {i['text']}" for i in all_items
                         if isinstance(chk.get(i['id']), dict) and chk[i['id']].get('status') == 'No')
        row_vals = [d.get('identity', {}).get('dept', ''), d.get('identity', {}).get('campus', ''),
                    d.get('identity', {}).get('semester', ''), yes, no, na, blank, pct, gaps]
        for c, v in enumerate(row_vals, 1):
            cell = ws6.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if c == 5 and isinstance(v, int) and v > 0:
                cell.fill = PatternFill("solid", fgColor="FFE8E8")
                cell.font = Font(name="Calibri", size=10, bold=True, color="AA0000")
        r += 1
    for i, w in enumerate([26, 12, 14, 8, 8, 8, 8, 14, 60], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    return wb



@app.route('/api/track-install', methods=['POST'])
def track_install():
    db['pwa_analytics'].update_one(
        {'_id': 'pwa_stats'},
        {'$inc': {'installs': 1}},
        upsert=True
    )
    return jsonify({'ok': True})

@app.route('/api/track-launch', methods=['POST'])
def track_launch():
    db['pwa_analytics'].update_one(
        {'_id': 'pwa_stats'},
        {'$inc': {'launches': 1}},
        upsert=True
    )
    return jsonify({'ok': True})

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs(os.path.join('static', 'uploads'), exist_ok=True)
    start_deadline_scheduler()
    port = int(os.environ.get('PORT', 5002))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(debug=debug, host='0.0.0.0', port=port)

